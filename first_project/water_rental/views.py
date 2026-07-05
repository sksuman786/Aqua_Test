from .models import SensorData, Customer, RequestLog
from django.utils import timezone

from datetime import timedelta, datetime
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import logging

logger = logging.getLogger(__name__)
from openpyxl import Workbook
from django.http import HttpResponse
from django.db import models
from django.db.models import Count
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.middleware.csrf import get_token
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone as django_timezone
from django.views.decorators.http import require_http_methods
from decimal import Decimal
import random
import json
from .models import (WaterQualityReading,
                     Customer, WaterUsage, Recharge,
                     Payment, UserComplain, Coupon,
                     CouponUsage, Technician,
                     SubscriptionPlan, CalibrationData,
                     EmailVerification,PasswordReset)
from django.urls import reverse
from django.views.decorators.http import require_http_methods
import time
import hmac
import hashlib
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.views.decorators.csrf import csrf_exempt
from .serializers import (
    SensorBatchSerializer, SensorDataSerializer, DailyConsumptionSerializer,
    MonthlyConsumptionSerializer, ProvisionDeviceSerializer, CheckOTASerializer,
    CustomerRegistrationSerializer, CustomerLoginSerializer, CustomerProfileSerializer,
    CustomerInfoSerializer, CustomerLocationSerializer, CustomerWaterUsageSerializer,
    CustomerRechargeInfoSerializer, CustomerProfileUpdateSerializer, ForgotPasswordRequestSerializer, 
    ForgotPasswordVerifySerializer, ForgotPasswordResetSerializer ,
    PaymentSerializer, PaymentCreateSerializer, SubscriptionInfoSerializer,
    UserComplainCustomerSerializer, UserComplainCreateSerializer,
    ESPBlockUnblockStatusSerializer, ESPOTAStatusCheckSerializer, ESPSensorDataSerializer,
    ESPDeviceHealthSerializer, ESPCalibrationFactorSerializer, FCMTokenSerializer
) 

from .utils import get_device_token
from .tasks import save_sensor_batch, save_sensor_data
from .totp_utils import generate_totp_secret, get_totp_code, verify_totp_code,send_totp_email, send_password_reset_email

import uuid
import razorpay
from django.conf import settings

from django.http import JsonResponse
from django.db.models import Count
from .models import RequestLog
import numpy as np

def device_count_api(request):
    data = (
        RequestLog.objects
        .values('source')
        .annotate(count=Count('source'))
        .order_by('-count')
    )

    return JsonResponse(list(data), safe=False)



import json
from django.db.models import Count
from django.shortcuts import render
from .models import RequestLog

_pending_password_resets = {}


def device_graph(request):
    qs = (
        RequestLog.objects
        .values('date', 'source')
        .annotate(count=Count('ip', distinct=True))
        .order_by('date')
    )

    sources = ["Browser", "Android", "ESP32", "iOS"]

    final = {}

    for row in qs:
        day = row['date']
        src = row['source']
        cnt = row['count']

        if day not in final:
            final[day] = {s: 0 for s in sources}
            final[day]['Total'] = 0

        if src in sources:
            final[day][src] = cnt
            final[day]['Total'] += cnt

    dates = [str(day) for day in final.keys()]
    browser = [final[d]["Browser"] for d in final]
    android = [final[d]["Android"] for d in final]
    esp = [final[d]["ESP32"] for d in final]
    ios = [final[d]["iOS"] for d in final]
    total = [final[d]["Total"] for d in final]

    return render(request, "chart.html", {
        "dates": json.dumps(dates),
        "browser": json.dumps(browser),
        "android": json.dumps(android),
        "esp": json.dumps(esp),
        "ios": json.dumps(ios),
        "total": json.dumps(total),
    })









"""=============start of the page views===================="""
"""when running the server redirect to admin login if not logged in and if logged in redirect to admin dashboard"""

def homepage(request):
    return render(request, 'landing_page.html')


def landing_page(request):
    """Landing page - redirects to appropriate page based on auth status"""
    # If user is authenticated, redirect to admin dashboard
    if request.user.is_authenticated:
        return redirect('aquaguard:admin-dashboard')
    # If not authenticated, redirect to login
    return redirect('aquaguard:admin-login')

"""Start of Admin  Views"""
"""================Adming Login/Signup Function============================="""

# Authentication for admin 

#check login credentials and authenticate admin


def admin_login(request):
    """Admin login view"""
    if request.user.is_authenticated:
        return redirect('aquaguard:admin-dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        try:
            # Get the first user with this email (in case of duplicates)
            user = User.objects.filter(email=email).first()
            if not user:
                messages.error(request, 'No account found with this email. Please sign up first.')
                return render(request, 'aquaguard/admin_login.html')
            
            authenticated_user = authenticate(request, username=user.username, password=password)
            
            if authenticated_user is not None:
                 # Check if this is the first login (last_login is None for new users)
                is_first_login = authenticated_user.last_login is None
                login(request, authenticated_user)
                
                # Handle Remember Me functionality
                if remember_me:
                    # Session expires after 30 days
                    request.session.set_expiry(2592000)  # 30 days in seconds
                else:
                    # Session expires when browser closes
                    request.session.set_expiry(0)

                #show welcome message
                if is_first_login:
                    messages.success(request, f'Welcome to AquaGuard Admin, {authenticated_user.username}! Your account has been created successfully. Please explore the dashboard to manage your customers and monitor water quality.')
                else:
                    messages.success(request, f'Welcome back to AquaGuard Water Quality Monitor, {authenticated_user.username}!')    
                
                              
                # Redirect to next parameter or admin dashboard
                next_url = request.GET.get('next', 'aquaguard:admin-dashboard')
                if next_url.startswith('/'):
                    return redirect(next_url)
                return redirect(next_url)
            else:
                messages.error(request, 'Incorrect password. Please try again.')
        except Exception as e:
            messages.error(request, f'Login error: {str(e)}')
    
    return render(request, 'aquaguard/admin_login.html')

"""================Admin Signup with OTP Verification============================="""

def admin_signup(request):
    """Admin signup view with OTP verification"""
    if request.user.is_authenticated:
        return redirect('aquaguard:admin-dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action', 'send_otp')
        
        if action == 'send_otp':
            # Step 1: Send OTP to email
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            password_confirm = request.POST.get('password_confirm')
            
            # Validation
            if not all([username, email, password, password_confirm]):
                messages.error(request, 'All fields are required.')
                return render(request, 'aquaguard/admin_signup.html')
            
            if password != password_confirm:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'aquaguard/admin_signup.html')
            
            if len(password) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return render(request, 'aquaguard/admin_signup.html')
            
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
                return render(request, 'aquaguard/admin_signup.html')
            
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already registered.')
                return render(request, 'aquaguard/admin_signup.html')
            
            # Generate 6-digit OTP
            otp = str(random.randint(100000, 999999))
            
            # Store data in session
            request.session['admin_signup_data'] = {
                'username': username,
                'email': email,
                'password': password,
                'otp': otp
            }
            
            # Send OTP via email
            try:
                send_mail(
                    subject='AquaGuard Admin - Email Verification OTP',
                    message=f'''Hello {username},

Your OTP for AquaGuard Admin signup is: {otp}

This OTP is valid for 10 minutes.

If you did not request this, please ignore this email.

Best regards,
AquaGuard Admin Team''',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
                messages.success(request, f"OTP has been sent to {email}. Please check your inbox.")
                return render(request, 'aquaguard/admin_signup.html', {'show_otp_field': True})
            except Exception as e:
                messages.error(request, f"Failed to send OTP: {str(e)}")
                return render(request, 'aquaguard/admin_signup.html')
        
        elif action == 'verify_otp':
            # Step 2: Verify OTP and create account
            entered_otp = request.POST.get('otp', '')
            try:
                signup_data = request.session.get('admin_signup_data')
            except Exception:
                messages.error(request, "Session expired or interrupted. Please start again.")
                return render(request, 'aquaguard/admin_signup.html')
            
            if not signup_data:
                messages.error(request, "Session expired. Please start again.")
                return render(request, 'aquaguard/admin_signup.html')
            
            if entered_otp != signup_data.get('otp'):
                messages.error(request, "Invalid OTP. Please try again.")
                return render(request, 'aquaguard/admin_signup.html', {'show_otp_field': True})
            
            # OTP verified, create account
            try:
                username = signup_data['username']
                email = signup_data['email']
                password = signup_data['password']
                
                # Create admin user
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    is_staff=True  # Make user staff
                )
                user.save()
                
                # Send credentials via email
                try:
                    send_mail(
                        subject='AquaGuard Admin - Account Created Successfully',
                        message=f'''Hello {username},

Your AquaGuard Admin account has been created successfully!

Your Admin Login Credentials:
━━━━━━━━━━━━━━━━━━━━━━━━
Username: {username}
Email: {email}
Password: {password}
━━━━━━━━━━━━━━━━━━━━━━━━

You can now login to your admin account at: http://127.0.0.1:8000/admin-login/

Please keep these credentials safe and change your password after first login.

Thank you for joining AquaGuard!

Best regards,
AquaGuard Admin Team''',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[email],
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Email sending failed: {e}")
                
                # Clear session data (guard against concurrent session deletion)
                try:
                    if 'admin_signup_data' in request.session:
                        del request.session['admin_signup_data']
                except Exception:
                    # If the session was deleted concurrently, ignore and continue
                    pass
                
                messages.success(request, 'Admin account created successfully! Login credentials have been sent to your email.')
                return redirect('aquaguard:admin-login')
            
            except Exception as e:
                messages.error(request, f'Error creating account: {str(e)}')
                return render(request, 'aquaguard/admin_signup.html')
    
    return render(request, 'aquaguard/admin_signup.html')

"""================Admin Logout Function============================="""

def admin_logout(request):
    """Admin logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('aquaguard:admin-login')


"""================Admin Dashboard Function============================="""

@login_required
def admin_dashboard(request):
    """Admin dashboard showing customer overview"""
    customers = Customer.objects.all()
    search_query = request.GET.get('search', '').strip()
    
    # Apply search filter
    if search_query:
        from django.db.models import Q
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(device_chip_id__icontains=search_query) |
            Q(customer_address__icontains=search_query)
        )
    
    # Enrich customer data with active recharge info and latest pending edit request
    from .models import ProfileEditRequest
    customer_list = []
    for customer in customers:
        active_recharge = Recharge.objects.filter(customer=customer, status='active').first()
        # Get latest pending edit request for this customer (if any)
        latest_edit_request = ProfileEditRequest.objects.filter(customer=customer, status='pending').order_by('-requested_at').first()
        # Use robust model property for price per litre
        price_per_litre = customer.paisa_per_litre
        customer_data = {
            'customer': customer,
            'active_recharge': active_recharge,
            'litres_remaining': active_recharge.litres_remaining if active_recharge else 0,
            'price_per_litre': price_per_litre,
            'edit_request_id': latest_edit_request.id if latest_edit_request else None,
        }
        customer_list.append(customer_data)
    
    context = {
        'total_customers': Customer.objects.all().count(),
        'online_devices': Customer.objects.filter(device_status='online').count(),
        'offline_devices': Customer.objects.filter(device_status='offline').count(),
        'customers': customer_list,
        'search_query': search_query,
    }
    return render(request, 'aquaguard/admin_dashboard.html', context)


"""================Customer Details View function============================="""

@login_required
def customer_details(request, customer_id):
    """Show detailed customer information with device and usage data"""
    from django.utils import timezone
    from django.db.models import Sum
    from datetime import datetime
    customer = Customer.objects.get(id=customer_id)
    usages = WaterUsage.objects.filter(customer=customer).order_by('-usage_date')
    recharges = Recharge.objects.filter(customer=customer).order_by('-recharge_date')
    payments = Payment.objects.filter(customer=customer).order_by('-recharge_date')
    
    # Get today's usage specifically for real-time display
    today = timezone.now().date()
    today_usage = WaterUsage.objects.filter(customer=customer, usage_date=today).first()

    usage_metrics = {
        'last_day_usage': 0.0,
        'today_usage': 0.0,
        'monthly_total_usage': 0.0,
        'monthly_average_usage': 0.0,
    }

    if today_usage and today_usage.current_day_usage is not None:
        usage_metrics['today_usage'] = float(today_usage.current_day_usage)

    yesterday = today - timedelta(days=1)
    yesterday_usage = WaterUsage.objects.filter(customer=customer, usage_date=yesterday).first()
    if yesterday_usage:
        daily_value = yesterday_usage.current_day_usage or yesterday_usage.last_24hr_usage or 0
        usage_metrics['last_day_usage'] = float(daily_value)

    month_window_start = today - timedelta(days=30)
    month_usages = WaterUsage.objects.filter(customer=customer, usage_date__gte=month_window_start, usage_date__lte=today)
    month_total = month_usages.aggregate(total=Sum('current_day_usage'))['total'] or 0
    month_total_value = float(month_total) if month_total else 0.0
    usage_metrics['monthly_total_usage'] = month_total_value
    if month_total_value:
        usage_metrics['monthly_average_usage'] = month_total_value / 30.0
    
    # Calculate totals across all active recharges
    active_recharges = recharges.filter(status='active')
    active_recharge_data = None
    if active_recharges.exists():
        totals = active_recharges.aggregate(
            total_allocated=Sum('litres_allocated'),
            total_used=Sum('litres_used'),
            total_remaining=Sum('litres_remaining')
        )
        # Get the earliest expiry date from active recharges
        earliest_expiry = active_recharges.order_by('expiry_date').first()
        active_recharge_data = {
            'litres_allocated': totals['total_allocated'],
            'litres_used': totals['total_used'],
            'litres_remaining': totals['total_remaining'],
            'days_until_expiry': earliest_expiry.days_until_expiry if earliest_expiry else 0,
        }
    
    # Get tracked location data
    tracked_location = None
    # Use only actual latitude/longitude/location fields
    latitude = customer.latitude
    longitude = customer.longitude
    location_name = customer.location or customer.customer_address or 'Unknown location'
    
    # Determine last updated time - use location update time if available, otherwise last recharge
    last_updated = customer.last_recharge_date or 'Not available'
    
    if latitude and longitude:
        tracked_location = {
            'latitude': latitude,
            'longitude': longitude,
            'location_name': location_name,
            'last_updated': last_updated,
            'has_resolved_location': bool(customer.location and customer.location != 'Unknown location')
        }
    
    from .models import ProfileEditRequest, CalibrationData
    latest_edit_request = ProfileEditRequest.objects.filter(customer=customer, status='pending').order_by('-requested_at').first()
    calibration_data = CalibrationData.objects.filter(customer=customer).order_by('-recorded_at')
    
    # Get latest sensor data for pH, TDS, water_quality, device_health
    from .models import SensorData
    latest_sensor = SensorData.objects.filter(customer=customer).order_by('-esp_timestamp').first()
    sensor_info = None
    if latest_sensor:
        sensor_info = {
            'ph': latest_sensor.ph,
            'tds': latest_sensor.tds,
            'water_quality': latest_sensor.water_quality,
            'device_health': latest_sensor.device_health,
            'flow_rate': latest_sensor.current_volume,
            'total_flow': latest_sensor.total_volume,
            'timestamp': latest_sensor.esp_timestamp,
        }
    
    context = {
        'customer': customer,
        'usages': usages,
        'recharges': recharges,
        'payments': payments,
        'last_usage': today_usage if today_usage else (usages.first() if usages.exists() else None),
        'usage_metrics': usage_metrics,
        'active_recharge': active_recharge_data,
        'active_recharge_breakdown': customer.paisa_per_litre_details,
        'tracked_location': tracked_location,
        'edit_request_id': latest_edit_request.id if latest_edit_request else None,
        'calibration_data': calibration_data,
        'customer_type_choices': Customer.CUSTOMER_TYPE_CHOICES,
        'board_options': BOARD_OPTIONS,
        'sensor_info': sensor_info,
    }
    return render(request, 'aquaguard/customer_details.html', context)

"""================Customer List View Class for in customer list============================="""


class CustomerListView(ListView):
    """List all customers"""
    model = Customer
    template_name = 'aquaguard/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20
    ordering = ['-registration_date']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        
        if search_query:
            # Search by name, email, phone, device ID, or address
            from django.db.models import Q
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone_number__icontains=search_query) |
                Q(device_chip_id__icontains=search_query) |
                Q(customer_address__icontains=search_query)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context

"""================Sensor Data Dashboard View function============================="""

@login_required
def sensor_data_dashboard(request):
    """Dashboard showing sensor data with usage calculations"""
    from .models import SensorData
    from django.db.models import Sum
    customers = Customer.objects.all()
    search_query = request.GET.get('search', '').strip()
    
    # Apply search filter
    if search_query:
        from django.db.models import Q
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(device_chip_id__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    customers = list(customers)
    customer_ids = [c.id for c in customers]
    sensor_data = []
    
    # Build a map of latest WaterQualityReading per customer
    water_quality_map = {}
    for wqr in WaterQualityReading.objects.order_by('-reading_date'):
        if wqr.customer_id and wqr.customer_id not in water_quality_map:
            water_quality_map[wqr.customer_id] = wqr
    today = timezone.now().date()

    recharge_map = {}
    if customer_ids:
        recharge_totals = Recharge.objects.filter(
            customer_id__in=customer_ids,
            status='active'
        ).values('customer_id').annotate(
            total_allocated=Sum('litres_allocated'),
            total_used=Sum('litres_used')
        )
        recharge_map = {entry['customer_id']: entry for entry in recharge_totals}

    for customer in customers:
        latest_usage = WaterUsage.objects.filter(customer=customer).order_by('-usage_date').first()
        # Get latest sensor reading for pH, TDS, water_quality, device_health
        latest_sensor = SensorData.objects.filter(customer=customer).order_by('-esp_timestamp').first()
        if latest_usage:
            recharge_stats = recharge_map.get(customer.id, {})
            total_allocated = float(recharge_stats.get('total_allocated') or 0)
            total_used = float(recharge_stats.get('total_used') or 0)
            remaining_water = max(total_allocated - total_used, 0)

            recent_readings = list(
                SensorData.objects.filter(customer=customer).order_by('-esp_timestamp')[:5]
            )
            latest_reading = recent_readings[0] if recent_readings else None

            # Use current_volume as flow_rate since it's the instantaneous measurement
            flow_rate_value = latest_reading.current_volume if latest_reading else None

            current_day_usage = 0.0
            if latest_usage.usage_date == today and latest_usage.current_day_usage is not None:
                current_day_usage = float(latest_usage.current_day_usage)
            sensor_data.append({
                'customer': customer,
                'device_status': customer.device_status,
                'last_24hr_usage': latest_usage.last_24hr_usage,
                'monthly_usage': latest_usage.monthly_usage,
                'per_day_average': latest_usage.per_day_average,
                'current_day_usage': current_day_usage,
                'allocated_water': total_allocated,
                'flow_rate': flow_rate_value,
                'total_used_water': total_used,
                'remaining_water': remaining_water,
                'ph': latest_sensor.ph if latest_sensor else None,
                'tds': latest_sensor.tds if latest_sensor else None,
                'water_quality': latest_sensor.water_quality if latest_sensor else None,
                'device_health': latest_sensor.device_health if latest_sensor else customer.device_health,
            })
    
    context = {
        'sensor_data': sensor_data,
        'search_query': search_query,
    }
    return render(request, 'aquaguard/sensor_data_dashboard.html', context)


"""================Payment History View function============================="""


@login_required
def payment_history(request):
    """Display payment history for all customers"""
    from django.db.models import Sum, Q
    
    payments = Payment.objects.all().order_by('-recharge_date')
    search_query = request.GET.get('search', '').strip()
    
    # Apply search filter
    if search_query:
        payments = payments.filter(
            Q(customer__name__icontains=search_query) |
            Q(customer__email__icontains=search_query) |
            Q(customer__phone_number__icontains=search_query) |
            Q(transaction_id__icontains=search_query) |
            Q(payment_status__icontains=search_query)
        )
    
    completed_payments = payments.filter(payment_status='completed')
    pending_payments = payments.filter(payment_status='pending')
    total_revenue = completed_payments.aggregate(Sum('amount'))['amount__sum'] or 0
    completed_count = completed_payments.count()
    pending_count = pending_payments.count()
    
    context = {
        'payments': payments,
        'total_revenue': total_revenue,
        'completed_count': completed_count,
        'pending_count': pending_count,
        'search_query': search_query,
    }
    return render(request, 'aquaguard/payment_history.html', context)

"""================Customer Search View function============================="""

@login_required
def customer_search(request):
    """Search customers by name, email, or phone number (case-insensitive)"""
    query = request.GET.get('query', '').strip()
    search_by = request.GET.get('search_by', 'name')
    
    customers = []
    search_type = ''
    customer_data = []
    
    if query:
        # Capitalize query for proper comparison
        query_capitalized = query.title()  # Capitalizes each word
        
        if search_by == 'name':
            # Case-insensitive search using icontains
            customers = Customer.objects.filter(name__icontains=query)
            search_type = 'Name'
        elif search_by == 'email':
            # Case-insensitive search for email
            customers = Customer.objects.filter(email__icontains=query)
            search_type = 'Email'
        elif search_by == 'phone':
            # Phone search doesn't need case handling
            customers = Customer.objects.filter(phone_number__icontains=query)
            search_type = 'Phone'
        
        # Get additional data for each customer
        for customer in customers:
            latest_usage = WaterUsage.objects.filter(customer=customer).first()
            latest_recharge = Recharge.objects.filter(customer=customer).order_by('-recharge_date').first()
            
            customer_data.append({
                'customer': customer,
                'last_24hr_usage': latest_usage.last_24hr_usage if latest_usage else 0,
                'last_recharge': latest_recharge,
            })
    
    context = {
        'customer_data': customer_data,
        'query': query,
        'search_by': search_by,
        'search_type': search_type,
    }
    return render(request, 'aquaguard/customer_search.html', context)


"""================Toggle Block/Unblock function============================="""


@login_required
@require_POST
def toggle_block_unblock(request, customer_id):
    """Toggle block/unblock status for a customer"""
    from django.core.mail import send_mail
    try:
        customer = Customer.objects.get(id=customer_id)
        old_status = customer.block_unblock
        customer.block_unblock = not customer.block_unblock
        customer.save()
        
        status = '🟢 Active' if customer.block_unblock else '🔴 Blocked'
        
        # Send notification to customer if blocked
        if not customer.block_unblock and old_status:
            try:
                send_mail(
                    'Your Water Machine is Blocked',
                    f'Dear {customer.name},\n\nYour water machine has been blocked by the administrator. Please contact support for more information.\n\nRegards,\nAquaGuard Team',
                    settings.DEFAULT_FROM_EMAIL,
                    [customer.email],
                    fail_silently=True
                )
            except Exception as e:
                logger.error(f"Failed to send block notification to {customer.email}: {e}")
        
        return JsonResponse({
            'success': True,
            'block_unblock': customer.block_unblock,
            'status': status,
            'message': f"Customer {customer.name} is now {status}"
        })
    except Customer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Customer not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)



"""================Business Alerts View function============================="""


@login_required
def business_alerts_view(request):
    """Display business alerts for admin"""
    from .business_alerts import get_all_business_alerts, get_alert_summary
    
    alerts = get_all_business_alerts()
    summary = get_alert_summary()
    search_query = request.GET.get('search', '').strip()
    
    # Apply search filter
    if search_query:
        search_lower = search_query.lower()
        alerts = [alert for alert in alerts if 
                 search_lower in alert.customer.name.lower() or
                 search_lower in alert.customer.email.lower() or
                 search_lower in alert.customer.phone_number.lower() or
                 search_lower in alert.alert_type.lower() or
                 search_lower in alert.message.lower() or
                 search_lower in alert.severity.lower()]
    
    context = {
        'alerts': alerts,
        'summary': summary,
        'search_query': search_query,
    }
    return render(request, 'aquaguard/business_alerts.html', context)


@login_required
def dismiss_alert(request, customer_id, alert_type):
    """Dismiss a specific alert"""
    try:
        from .models import DismissedAlert
        
        customer = Customer.objects.get(id=customer_id)
        
        # Create or update dismissed alert
        dismissed, created = DismissedAlert.objects.get_or_create(
            customer=customer,
            alert_type=alert_type,
            defaults={'dismissed_by': request.user}
        )
        
        return JsonResponse({
            'status': 'success',
            'message': f'Alert dismissed for {customer.name}',
            'alert_type': alert_type
        })
    except Customer.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Customer not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Database not ready. Please run migrations first: {str(e)}'
        }, status=500)


@login_required
def restore_alert(request, customer_id, alert_type):
    """Restore a dismissed alert"""
    try:
        from .models import DismissedAlert
        
        customer = Customer.objects.get(id=customer_id)
        DismissedAlert.objects.filter(
            customer=customer,
            alert_type=alert_type
        ).delete()
        
        return JsonResponse({
            'status': 'success',
            'message': f'Alert restored for {customer.name}'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
def get_alerts_json(request):
    """API endpoint to get alerts as JSON for real-time updates"""
    from .business_alerts import get_all_business_alerts, get_alert_summary
    
    alerts = get_all_business_alerts()
    summary = get_alert_summary()
    
    alerts_data = []
    for alert in alerts:
        alerts_data.append({
            'customer_id': alert.customer.id,
            'customer_name': alert.customer.name,
            'alert_type': alert.alert_type,
            'severity': alert.severity,
            'message': alert.message,
            'details': alert.details,
        })
    
    return JsonResponse({
        'alerts': alerts_data,
        'summary': summary,
        'timestamp': django_timezone.now().isoformat()
    })





# ========== TECHNICIAN MANAGEMENT VIEWS ========== 

BOARD_OPTIONS = [('esp32', 'ESP32'), ('esp8266', 'ESP8266')]
BOARD_OPTION_VALUES = {value for value, _ in BOARD_OPTIONS}
@login_required
def technician_list(request):
    technicians = Technician.objects.all()
    context = {'technicians': technicians}
    return render(request, 'aquaguard/technician_list.html', context)

@login_required
@require_http_methods(["GET", "POST"])
def add_technician(request):
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        specialization = request.POST.get('specialization')
        area_of_operation = request.POST.get('area_of_operation')
        rating = request.POST.get('rating') or 0
        completed_jobs = request.POST.get('completed_jobs') or 0
        status = request.POST.get('status') or 'available'
        is_active = bool(request.POST.get('is_active'))
        Technician.objects.create(
            name=name,
            email=email,
            phone_number=phone_number,
            specialization=specialization,
            area_of_operation=area_of_operation,
            rating=rating,
            completed_jobs=completed_jobs,
            status=status,
            is_active=is_active
        )
        return redirect('aquaguard:technician-list')
    return render(request, 'aquaguard/add_technician.html')

#edit technician details

@login_required
@require_http_methods(["GET", "POST"])
def edit_technician(request, technician_id):
    technician = Technician.objects.get(id=technician_id)
    if request.method == "POST":
        technician.name = request.POST.get('name')
        technician.email = request.POST.get('email')
        technician.phone_number = request.POST.get('phone_number')
        technician.specialization = request.POST.get('specialization')
        technician.area_of_operation = request.POST.get('area_of_operation')
        technician.rating = request.POST.get('rating') or 0
        technician.completed_jobs = request.POST.get('completed_jobs') or 0
        technician.status = request.POST.get('status') or 'available'
        technician.is_active = bool(request.POST.get('is_active'))
        technician.save()
        return redirect('aquaguard:technician-list')
    context = {'technician': technician}
    return render(request, 'aquaguard/edit_technician.html', context)

#delete technician

@login_required
def delete_technician(request, technician_id):
    technician = Technician.objects.get(id=technician_id)
    technician.delete()
    return redirect('aquaguard:technician-list')

#view technician details

@login_required
def technician_detail(request, technician_id):
    technician = Technician.objects.get(id=technician_id)
    context = {'technician': technician}
    return render(request, 'aquaguard/technician_detail.html', context)


# ========== SUBSCRIPTION PLAN MANAGEMENT ==========

@login_required
def admin_subscription_plans(request):
    """List all subscription plans with filters"""
    plans = SubscriptionPlan.objects.all().order_by('-created_at')
    
    # Filters
    status_filter = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    if status_filter:
        plans = plans.filter(status=status_filter)
    if search:
        plans = plans.filter(plan_name__icontains=search)
    
    context = {
        'plans': plans,
        'current_status': status_filter,
        'current_search': search,
    }
    return render(request, 'aquaguard/admin_subscription_plans.html', context)

#================Subscription Plan Detail View function============================="""

@login_required
def admin_subscription_plan_detail(request, plan_id):
    """View subscription plan details"""
    plan = SubscriptionPlan.objects.get(id=plan_id)
    
    # Get customers using this plan
    customers = Customer.objects.filter(subscription_plan=plan).select_related('subscription_plan')
    
    context = {
        'plan': plan,
        'customers': customers,
        'customer_count': customers.count(),
        'price_per_litre': plan.paisa_per_litre,
    }
    return render(request, 'aquaguard/admin_subscription_plan_detail.html', context)


#================Create Subscription Plan View function============================="""

@login_required
def admin_create_subscription_plan(request):
    """Create a new subscription plan"""
    if request.method == 'POST':
        try:
            plan_name = request.POST.get('plan_name', '').strip()
            description = request.POST.get('description', '').strip()
            price = request.POST.get('price', '')
            litres_allocated = request.POST.get('litres_allocated', '')
            duration_days = request.POST.get('duration_days', '30')
            is_popular = request.POST.get('is_popular') == 'on'
            status = request.POST.get('status') == 'true'
            features = request.POST.getlist('features[]')
            
            # Validation
            errors = []
            if not plan_name:
                errors.append('Plan name is required')
            if not price:
                errors.append('Price is required')
            if not litres_allocated:
                errors.append('Litres allocated is required')
            
            try:
                price = float(price)
                litres_allocated = int(litres_allocated)
                duration_days = int(duration_days)
            except ValueError:
                errors.append('Price must be a number, litres and days must be integers')
            
            if price <= 0:
                errors.append('Price must be greater than 0')
            if litres_allocated <= 0:
                errors.append('Litres allocated must be greater than 0')
            if duration_days <= 0:
                errors.append('Duration must be greater than 0')
            
            if errors:
                context = {
                    'errors': errors,
                    'form_data': request.POST,
                }
                return render(request, 'aquaguard/admin_create_subscription_plan.html', context)
            
            # Create plan
            plan = SubscriptionPlan.objects.create(
                plan_name=plan_name,
                description=description,
                price=price,
                litres_allocated=litres_allocated,
                duration_days=duration_days,
                is_popular=is_popular,
                status=status,
                description_features=features if features else [],
            )
            
            messages.success(request, f'✅ Subscription plan "{plan_name}" created successfully!')
            return redirect('aquaguard:subscription-plans')
            
        except Exception as e:
            messages.error(request, f'❌ Error creating plan: {str(e)}')
            return redirect('aquaguard:subscription-plans')
    
    context = {
    }
    return render(request, 'aquaguard/admin_create_subscription_plan.html', context)


#================Update Subscription Plan View function============================="""

@login_required
def admin_update_subscription_plan(request, plan_id):
    """Update an existing subscription plan"""
    plan = SubscriptionPlan.objects.get(id=plan_id)
    
    if request.method == 'POST':
        try:
            plan_name = request.POST.get('plan_name', '').strip()
            description = request.POST.get('description', '').strip()
            price = request.POST.get('price', '')
            litres_allocated = request.POST.get('litres_allocated', '')
            duration_days = request.POST.get('duration_days', '30')
            is_popular = request.POST.get('is_popular') == 'on'
            status = request.POST.get('status') == 'true'
            features = request.POST.getlist('features[]')
            
            # Validation
            errors = []
            if not plan_name:
                errors.append('Plan name is required')
            if not price:
                errors.append('Price is required')
            if not litres_allocated:
                errors.append('Litres allocated is required')
            
            try:
                price = float(price)
                litres_allocated = int(litres_allocated)
                duration_days = int(duration_days)
            except ValueError:
                errors.append('Price must be a number, litres and days must be integers')
            
            if price <= 0:
                errors.append('Price must be greater than 0')
            if litres_allocated <= 0:
                errors.append('Litres allocated must be greater than 0')
            if duration_days <= 0:
                errors.append('Duration must be greater than 0')
            
            if errors:
                context = {
                    'plan': plan,
                    'errors': errors,
                    'form_data': request.POST,
                }
                return render(request, 'aquaguard/admin_update_subscription_plan.html', context)
            
            # Update plan
            plan.plan_name = plan_name
            plan.description = description
            plan.price = price
            plan.litres_allocated = litres_allocated
            plan.duration_days = duration_days
            plan.is_popular = is_popular
            plan.status = status
            plan.description_features = features if features else []
            plan.save()
            
            messages.success(request, f'✅ Subscription plan "{plan_name}" updated successfully!')
            return redirect('aquaguard:subscription-plan-detail', plan_id=plan.id)
            
        except Exception as e:
            messages.error(request, f'❌ Error updating plan: {str(e)}')
            return redirect('aquaguard:subscription-plan-detail', plan_id=plan.id)
    
    context = {
        'plan': plan,
    }
    return render(request, 'aquaguard/admin_update_subscription_plan.html', context)


#================Delete Subscription Plan View function============================="""

@login_required
@require_http_methods(["POST"])
def admin_delete_subscription_plan(request, plan_id):
    """Delete a subscription plan"""
    try:
        plan = SubscriptionPlan.objects.get(id=plan_id)
        plan_name = plan.plan_name
        
        # Check if any customers are using this plan
        customer_count = plan.customer_set.count()
        if customer_count > 0:
            messages.warning(request, f'⚠️ Cannot delete plan "{plan_name}" - {customer_count} customer(s) are using it. Remove customers first.')
            return redirect('aquaguard:subscription-plan-detail', plan_id=plan.id)
        
        plan.delete()
        messages.success(request, f'✅ Subscription plan "{plan_name}" deleted successfully!')
        return redirect('aquaguard:subscription-plans')
        
    except SubscriptionPlan.DoesNotExist:
        messages.error(request, '❌ Plan not found')
        return redirect('aquaguard:subscription-plans')
    except Exception as e:
        messages.error(request, f'❌ Error deleting plan: {str(e)}')
        return redirect('aquaguard:subscription-plans')
    

#================User Complaints Management Views============================="""


@login_required
def admin_complaints_list(request):
    """Display all user complaints with filtering and search"""
    complaints = UserComplain.objects.all().order_by('-complain_date')
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    overdue_filter = request.GET.get('overdue', '')
    
    if search_query:
        from django.db.models import Q
        complaints = complaints.filter(
            Q(customer__name__icontains=search_query) |
            Q(phone_no__icontains=search_query) |
            Q(device_id__icontains=search_query) |
            Q(problem_statement__icontains=search_query) |
            Q(future_field1__icontains=search_query)
        )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    if priority_filter:
        complaints = complaints.filter(priority=priority_filter)
    
    if overdue_filter:
        complaints = complaints.filter(
            status__in=['open', 'pending'],
            complain_date__lt=django_timezone.now() - django_timezone.timedelta(days=7)
        )
    
    # Calculate counts
    all_complaints = UserComplain.objects.all()
    open_count = all_complaints.filter(status='open').count()
    in_progress_count = all_complaints.filter(status='in_progress').count()
    resolved_count = all_complaints.filter(status='resolved').count()
    closed_count = all_complaints.filter(status='closed').count()
    overdue_count = all_complaints.filter(
        status__in=['open', 'pending'],
        complain_date__lt=django_timezone.now() - django_timezone.timedelta(days=7)
    ).count()
    
    # Ensure future_field1 and future_field2 are set for display
    for complaint in complaints:
        # Skip data migration since future_field1 is now the primary email storage
        if complaint.closed_date and (not complaint.future_field2 or complaint.future_field2 == ''):
            complaint.future_field2 = complaint.closed_date.strftime('%Y-%m-%d %H:%M')
            complaint.save(update_fields=['future_field2'])

    context = {
        'complaints': complaints,
        'total_complaints': all_complaints.count(),
        'open_count': open_count,
        'in_progress_count': in_progress_count,
        'resolved_count': resolved_count,
        'closed_count': closed_count,
        'overdue_count': overdue_count,
        'search_query': search_query,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
    }
    return render(request, 'aquaguard/admin_complaints_list.html', context)

#================User Complaint Detail View function============================="""


@login_required
def admin_complaint_detail(request, complaint_id):
    """Display and manage a specific complaint"""
    complaint = UserComplain.objects.get(id=complaint_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'assign':
            assigned_technician_id = request.POST.get('assigned_to')
            if assigned_technician_id:
                from .models import Technician
                assigned_technician = Technician.objects.get(id=assigned_technician_id)
                complaint.assigned_person = assigned_technician
                complaint.assigned_person_details = request.POST.get('notes', '')
                complaint.assigned_at = django_timezone.now()
                complaint.status = 'in_progress'
                complaint.save()
                messages.success(request, f'Complaint assigned to {assigned_technician.name}')

        elif action == 'resolve':
            complaint.status = 'resolved'
            complaint.resolution_notes = request.POST.get('notes', '')
            complaint.save()
            messages.success(request, 'Complaint marked as resolved')

        elif action == 'close':
            complaint.status = 'closed'
            complaint.closed_date = django_timezone.now()
            # Append closing note to resolution notes for history
            closing_note = request.POST.get('notes', '')
            if closing_note:
                existing = complaint.resolution_notes or ''
                complaint.resolution_notes = (existing + '\n' if existing else '') + closing_note
            complaint.save()
            messages.success(request, 'Complaint closed successfully')

        return redirect('aquaguard:admin-complaint-detail', complaint_id=complaint_id)

    from .models import Technician
    technicians = Technician.objects.all()

    context = {
        'complaint': complaint,
        'technicians': technicians,
    }
    return render(request, 'aquaguard/admin_complaint_detail.html', context)


#================User Complaints Dashboard View function============================="""

@login_required
def admin_complaints_dashboard(request):
    """Dashboard view for complaint statistics and analytics"""
    all_complaints = UserComplain.objects.all()
    total_complaints = all_complaints.count()
    
    open_count = all_complaints.filter(status='open').count()
    in_progress_count = all_complaints.filter(status='in_progress').count()
    resolved_count = all_complaints.filter(status='resolved').count()
    closed_count = all_complaints.filter(status='closed').count()
    
    status_breakdown = {
        'open': open_count,
        'in_progress': in_progress_count,
        'resolved': resolved_count,
        'closed': closed_count,
        'pending': all_complaints.filter(status='pending').count(),
    }
    
    priority_breakdown = {
        'low': all_complaints.filter(priority='low').count(),
        'medium': all_complaints.filter(priority='medium').count(),
        'high': all_complaints.filter(priority='high').count(),
        'critical': all_complaints.filter(priority='critical').count(),
    }
    
    overdue_complaints = all_complaints.filter(
        status__in=['open', 'pending'],
        complain_date__lt=django_timezone.now() - django_timezone.timedelta(days=7)
    )
    
    recent_complaints = all_complaints.order_by('-complain_date')[:10]
    
    # Calculate resolution rate
    if total_complaints > 0:
        resolution_rate = int((closed_count / total_complaints) * 100)
    else:
        resolution_rate = 0
    
    context = {
        'total_complaints': total_complaints,
        'open_count': open_count,
        'in_progress_count': in_progress_count,
        'resolved_count': resolved_count,
        'closed_count': closed_count,
        'status_breakdown': status_breakdown,
        'priority_breakdown': priority_breakdown,
        'overdue_count': overdue_complaints.count(),
        'overdue_complaints': overdue_complaints,
        'recent_complaints': recent_complaints,
        'resolution_rate': resolution_rate,
    }
    return render(request, 'aquaguard/admin_complaints_dashboard.html', context)


#================Create User Complaint View function============================="""

@login_required
def admin_create_complaint(request):
    """Create a new complaint manually"""
    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        phone_no = request.POST.get('phone_no', '').strip()
        device_id = request.POST.get('device_id', '').strip()
        problem_statement = request.POST.get('problem_statement', '').strip()
        priority = request.POST.get('priority', 'medium')
        status = request.POST.get('status', 'open')
        assigned_person_id = request.POST.get('assigned_person')

        # Validation
        if not customer_id:
            messages.error(request, 'Please select a customer')
            return redirect('aquaguard:admin-create-complaint')

        if not problem_statement:
            messages.error(request, 'Please provide a problem description')
            return redirect('aquaguard:admin-create-complaint')

        try:
            customer = Customer.objects.get(id=customer_id)
        except Customer.DoesNotExist:
            messages.error(request, 'Selected customer not found')
            return redirect('aquaguard:admin-create-complaint')

        # Create the complaint
        complaint = UserComplain.objects.create(
            customer=customer,
            phone_no=phone_no or customer.phone_number,
            device_id=device_id or customer.device_chip_id,
            problem_statement=problem_statement,
            priority=priority,
            status=status,
            future_field1=customer.email,  # Store email in future_field1
        )

        # Assign technician if selected
        if assigned_person_id:
            try:
                from .models import Technician
                technician = Technician.objects.get(id=assigned_person_id)
                complaint.assigned_person = technician
                complaint.assigned_at = django_timezone.now()
                if status == 'open':
                    complaint.status = 'in_progress'
                complaint.save()
            except Technician.DoesNotExist:
                pass  # Continue without assignment

        messages.success(request, f'Complaint created successfully! Ticket #{complaint.ticket_number}')
        return redirect('aquaguard:admin-complaint-detail', complaint_id=complaint.id)

    # GET request - show form
    from .models import Technician
    technicians = Technician.objects.filter(is_active=True)

    context = {
        'technicians': technicians,
    }
    return render(request, 'aquaguard/admin_create_complaint.html', context)



"""================Export Customers Data to Excel View function============================="""

@login_required
def export_customers_excel(request):
    """Export all customer details to Excel file (.xlsx)"""
    from datetime import datetime
    customers = Customer.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    # Header row
    ws.append([
        'Customer ID', 'Name', 'Email', 'Phone', 'Device ID', 'Current Plan', 'Current Plan Price',
        'Last Recharge Date', 'Registration Date', 'Status', 'Location', 'Customer Address',
        'Latitude', 'Longitude', 'Block Status'
    ])
    # Data rows
    for customer in customers:
        plan_name = customer.subscription_plan.plan_name if customer.subscription_plan else ''
        ws.append([
            str(customer.id),
            customer.name,
            customer.email,
            customer.phone_number,
            customer.device_chip_id,
            plan_name,
            customer.paisa_per_litre,
            customer.last_recharge_date.strftime('%Y-%m-%d %H:%M:%S') if customer.last_recharge_date else '',
            customer.registration_date.strftime('%Y-%m-%d %H:%M:%S') if customer.registration_date else '',
            'Active' if customer.block_unblock else 'Blocked',
            customer.location if customer.location else '',
            customer.customer_address if customer.customer_address else '',
            customer.latitude if customer.latitude else '',
            customer.longitude if customer.longitude else '',
            'No' if customer.block_unblock else 'Yes'
        ])
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

"""================Export Sensor Data to CSV View function============================="""

@login_required
def export_sensor_data_csv(request):
    """Export all sensor/water quality data to CSV file"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    try:
        from .models import SensorData, Customer
        from django.utils import timezone
        from datetime import timedelta
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.utils import get_column_letter

        # Get all per-second sensor data
        readings = SensorData.objects.select_related('customer').all().order_by('-esp_timestamp')
        # Build a map of latest WaterQualityReading per customer (keyed by customer UUID)
        water_quality_map = {}
        for wqr in WaterQualityReading.objects.order_by('-reading_date'):
            if wqr.customer_id and wqr.customer_id not in water_quality_map:
                water_quality_map[wqr.customer_id] = wqr

        # Precompute per-customer usage for last 24hr, week, month, year using WaterUsage model
        now = timezone.now().date()
        usage_24hr = {}
        usage_week = {}
        usage_month = {}
        usage_year = {}
        from .models import WaterUsage
        for customer in Customer.objects.all():
            cid = customer.email
            # Last 24hr usage: get the latest WaterUsage for today
            wu_24hr = WaterUsage.objects.filter(customer=customer, usage_date=now).order_by('-usage_date').first()
            usage_24hr[cid] = wu_24hr.last_24hr_usage if wu_24hr else 0
            # Last 7 days usage: sum last 7 days
            wu_week = WaterUsage.objects.filter(customer=customer, usage_date__gte=now - timedelta(days=6), usage_date__lte=now)
            usage_week[cid] = sum(wu.last_24hr_usage for wu in wu_week) if wu_week else 0
            # Last 30 days usage: sum last 30 days
            wu_month = WaterUsage.objects.filter(customer=customer, usage_date__gte=now - timedelta(days=29), usage_date__lte=now)
            usage_month[cid] = sum(wu.last_24hr_usage for wu in wu_month) if wu_month else 0
            # Last 365 days usage: sum last 365 days
            wu_year = WaterUsage.objects.filter(customer=customer, usage_date__gte=now - timedelta(days=364), usage_date__lte=now)
            usage_year[cid] = sum(wu.last_24hr_usage for wu in wu_year) if wu_year else 0

        # Create Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sensor Data"

        # Write header row
        headers = [
            'ID', 'Customer Name', 'Customer Email', 'Device Chip ID', 'Total Flow (L)',
            'pH', 'TDS (ppm)', 'Water Quality', 'Timestamp', 'Date',
            'Litres Used (24hr)', 'Litres Used (Week)', 'Litres Used (Month)', 'Litres Used (Year)'
        ]
        ws.append(headers)

        # Write per-second sensor data rows (with per-reading increment)
        # Build a dict to track previous total_flow for each customer per day
        prev_total_flow = {}
        for reading in readings:
            customer_email = reading.customer.email if reading.customer else None
            customer_uuid = reading.customer.id if reading.customer else None
            date_key = (customer_email, reading.esp_timestamp.date())
            prev = prev_total_flow.get(date_key, None)
            if prev is not None:
                increment = reading.total_volume - prev
                if increment < 0:
                    increment = 0
            else:
                increment = reading.total_volume
            prev_total_flow[date_key] = reading.total_volume
            import pytz
            kolkata = pytz.timezone('Asia/Kolkata')
            local_timestamp = reading.esp_timestamp.astimezone(kolkata) if reading.esp_timestamp else None
            # Use pH, TDS, water_quality directly from SensorData reading
            ws.append([
                reading.id,
                reading.customer.name if reading.customer else '',
                customer_email or '',
                reading.customer.device_chip_id if reading.customer else '',
                reading.total_volume,
                reading.ph if reading.ph is not None else '',
                reading.tds if reading.tds is not None else '',
                reading.water_quality or '',
                local_timestamp.strftime('%Y-%m-%d %H:%M:%S') if local_timestamp else '',
                reading.esp_timestamp.date().strftime('%Y-%m-%d') if reading.esp_timestamp else '',
                round(reading.total_volume, 3),
            ])
        # Optionally, update headers to reflect the change
        ws.delete_rows(1)
        headers = [
            'ID', 'Customer Name', 'Customer Email', 'Device Chip ID', 'Total Flow (L)',
            'pH', 'TDS (ppm)', 'Water Quality', 'Timestamp', 'Date',
            'Total Water Consumed (L)'
        ]
        ws.insert_rows(1)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Create HTTP response with Excel content type
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"sensor_data_export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        messages.success(request, f"Exported {readings.count()} per-second sensor records with per-customer usage stats (Excel)")
        return response

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting per-second sensor data to Excel: {str(e)}")
        messages.error(request, "Error exporting per-second sensor data")
        return redirect('aquaguard:reading-list')


"""================Export Water Usage Data to Excel View function============================="""

@login_required
def export_water_usage_excel(request):
    """Export all water usage data to Excel file (.xlsx)"""
    from datetime import datetime
    
    try:
        # Get all water usage records
        usage_records = WaterUsage.objects.select_related('customer').all().order_by('-usage_date')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Water Usage"
        
        # Write header row
        ws.append([
            'Customer Name',
            'Customer Email',
            'Usage Date',
            'Last 24hr Usage',
            'Monthly Usage',
            'Per Day Average',
            'Current Day Usage'
        ])
        
        # Write water usage rows
        for usage in usage_records:
            ws.append([
                usage.customer.name,
                usage.customer.email,
                usage.usage_date.strftime('%b. %d, %Y') if usage.usage_date else '',
                float(usage.last_24hr_usage) if usage.last_24hr_usage else 0,
                float(usage.monthly_usage) if usage.monthly_usage else 0,
                float(usage.per_day_average) if usage.per_day_average else 0,
                float(usage.current_day_usage) if usage.current_day_usage else 0
            ])
        
        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"water_usage_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        messages.success(request, f"Exported {usage_records.count()} water usage records to Excel")
        return response
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting water usage to Excel: {str(e)}")
        messages.error(request, "Error exporting water usage data")
        return redirect('aquaguard:dashboard')


# ========== COUPON MANAGEMENT ==========

#================Coupons List View function============================="""

@login_required
def admin_coupons_list(request):
    """List all coupons with filters"""
    coupons = Coupon.objects.all().order_by('-created_at')
    
    # Filters
    status = request.GET.get('status', '')
    coupon_type = request.GET.get('type', '')
    search = request.GET.get('search', '')
    
    if status:
        coupons = coupons.filter(status=status)
    if coupon_type:
        coupons = coupons.filter(coupon_type=coupon_type)
    if search:
        coupons = coupons.filter(coupon_code__icontains=search)
    
    context = {
        'coupons': coupons,
        'status_choices': Coupon.COUPON_STATUS_CHOICES,
        'type_choices': Coupon.COUPON_TYPE_CHOICES,
        'current_status': status,
        'current_type': coupon_type,
        'current_search': search,
        'future_field1': getattr(Coupon, 'future_field1', None),
        'future_field2': getattr(Coupon, 'future_field2', None),
    }
    return render(request, 'aquaguard/admin_coupons_list.html', context)


#================Create Coupon View function============================="""

@login_required
def admin_create_coupon(request):
    """Create new coupon"""
    from datetime import datetime, timedelta, timezone
    from django.http import JsonResponse
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validation
            coupon_code = data.get('coupon_code', '').strip().upper()
            if not coupon_code:
                return JsonResponse({'success': False, 'error': 'Coupon code is required'})
            
            # Check if code already exists
            if Coupon.objects.filter(coupon_code=coupon_code).exists():
                return JsonResponse({'success': False, 'error': 'Coupon code already exists'})
            
            coupon_type = data.get('coupon_type', 'percentage')
            if coupon_type not in ['percentage', 'fixed', 'free_litres']:
                return JsonResponse({'success': False, 'error': 'Invalid coupon type'})
            
            discount_value = data.get('discount_value')
            if not discount_value:
                return JsonResponse({'success': False, 'error': 'Discount value is required'})
            
            try:
                discount_value = Decimal(str(discount_value))
                if discount_value <= 0:
                    return JsonResponse({'success': False, 'error': 'Discount value must be greater than 0'})
            except:
                return JsonResponse({'success': False, 'error': 'Invalid discount value'})
            
            valid_from = data.get('valid_from')
            valid_until = data.get('valid_until')
            if not valid_from or not valid_until:
                return JsonResponse({'success': False, 'error': 'Valid dates are required'})
            
            try:
                valid_from = datetime.fromisoformat(valid_from.replace('Z', '+00:00'))
                valid_until = datetime.fromisoformat(valid_until.replace('Z', '+00:00'))
            except:
                return JsonResponse({'success': False, 'error': 'Invalid date format'})
            
            if valid_from >= valid_until:
                return JsonResponse({'success': False, 'error': 'Valid from date must be before valid until date'})
            
            # Create coupon
            coupon = Coupon.objects.create(
                coupon_code=coupon_code,
                coupon_type=coupon_type,
                discount_value=discount_value,
                valid_from=valid_from,
                valid_until=valid_until,
                description=data.get('description', ''),
                max_usage=data.get('max_usage') or None,
                max_usage_per_customer=int(data.get('max_usage_per_customer', 1)),
                min_order_amount=Decimal(str(data.get('min_order_amount', 0))),
                status=data.get('status', 'active') == 'active',
            )
            
            # Add applicable plans if specific plans selected
            apply_to = data.get('apply_to', 'all_plans')
            if apply_to == 'specific_plans':
                plan_ids = data.get('applicable_plans', [])
                if plan_ids:
                    coupon.applicable_plans.set(plan_ids)
            
            return JsonResponse({
                'success': True,
                'message': 'Coupon created successfully',
                'coupon_id': coupon.id
            })
        
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating coupon: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    from datetime import datetime, timedelta, timezone
    plans = SubscriptionPlan.objects.filter(status=True)
    context = {
        'plans': plans,
        'type_choices': Coupon.COUPON_TYPE_CHOICES,
        'status_choices': Coupon.COUPON_STATUS_CHOICES,
        'default_from': datetime.now(timezone.utc).isoformat(),
        'default_until': (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
    }
    return render(request, 'aquaguard/admin_create_coupon.html', context)


#================Update Coupon View function============================="""

@login_required
def admin_update_coupon(request, coupon_id):
    """Update existing coupon"""
    from django.http import JsonResponse
    import json
    from decimal import Decimal
    from datetime import datetime
    
    try:
        coupon = Coupon.objects.get(id=coupon_id)
    except Coupon.DoesNotExist:
        if request.method == 'POST':
            return JsonResponse({'success': False, 'error': 'Coupon not found'})
        messages.error(request, 'Coupon not found')
        return redirect('aquaguard:admin-coupons')
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Update basic fields
            coupon.coupon_type = data.get('coupon_type', coupon.coupon_type)
            coupon.description = data.get('description', coupon.description)
            coupon.status = data.get('status', coupon.status) == 'active'
            
            discount_value = data.get('discount_value')
            if discount_value:
                coupon.discount_value = Decimal(str(discount_value))
            
            valid_from = data.get('valid_from')
            valid_until = data.get('valid_until')
            if valid_from:
                coupon.valid_from = datetime.fromisoformat(valid_from.replace('Z', '+00:00'))
            if valid_until:
                coupon.valid_until = datetime.fromisoformat(valid_until.replace('Z', '+00:00'))
            
            coupon.max_usage = data.get('max_usage') or None
            coupon.max_usage_per_customer = int(data.get('max_usage_per_customer', 1))
            coupon.min_order_amount = Decimal(str(data.get('min_order_amount', 0)))
            
            coupon.save()
            
            # Handle applicable plans
            apply_to = data.get('apply_to', 'all_plans')
            if apply_to == 'specific_plans':
                plan_ids = data.get('applicable_plans', [])
                coupon.applicable_plans.set(plan_ids)
            else:
                coupon.applicable_plans.clear()
            
            return JsonResponse({
                'success': True,
                'message': 'Coupon updated successfully'
            })
        
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error updating coupon: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    plans = SubscriptionPlan.objects.filter(status=True)
    coupon_plan_ids = list(coupon.applicable_plans.values_list('id', flat=True))
    context = {
        'coupon': coupon,
        'plans': plans,
        'coupon_plan_ids': coupon_plan_ids,
        'type_choices': Coupon.COUPON_TYPE_CHOICES,
        'status_choices': Coupon.COUPON_STATUS_CHOICES,
    }
    return render(request, 'aquaguard/admin_update_coupon.html', context)


#================Delete Coupon View function============================="""

@login_required
def admin_delete_coupon(request, coupon_id):
    """Delete coupon"""
    from django.http import JsonResponse
    import json
    
    try:
        coupon = Coupon.objects.get(id=coupon_id)
    except Coupon.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Coupon not found'})
    
    if request.method == 'POST':
        try:
            coupon_code = coupon.coupon_code
            coupon.delete()
            return JsonResponse({
                'success': True,
                'message': f'Coupon {coupon_code} deleted successfully'
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error deleting coupon: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


#================Coupon Detail View function============================="""

@login_required
def admin_coupon_detail(request, coupon_id):
    """View coupon details and usage statistics"""
    try:
        coupon = Coupon.objects.get(id=coupon_id)
    except Coupon.DoesNotExist:
        messages.error(request, 'Coupon not found')
        return redirect('aquaguard:admin-coupons')
    
    from .models import CouponUsage
    
    usages = CouponUsage.objects.filter(coupon=coupon).select_related('customer').order_by('-used_at')
    
    context = {
        'coupon': coupon,
        'usages': usages,
        'usage_count': usages.count(),
        'total_discount': sum(usage.discount_amount for usage in usages),
    }
    return render(request, 'aquaguard/admin_coupon_detail.html', context)


#================Motivational Quotes Management Views============================="""

@login_required
def admin_quotes_management(request):
    """Admin view to manage motivational water quotes"""
    from .models import MotivationalQuote
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle':
            # Toggle quote active status
            quote_id = request.POST.get('quote_id')
            try:
                quote = MotivationalQuote.objects.get(id=quote_id)
                quote.is_active = not quote.is_active
                quote.save()
                status = 'activated' if quote.is_active else 'deactivated'
                messages.success(request, f'Quote {status} successfully!')
            except MotivationalQuote.DoesNotExist:
                messages.error(request, 'Quote not found')
        
        elif action == 'delete':
            # Delete quote
            quote_id = request.POST.get('quote_id')
            try:
                quote = MotivationalQuote.objects.get(id=quote_id)
                quote.delete()
                messages.success(request, 'Quote deleted successfully!')
            except MotivationalQuote.DoesNotExist:
                messages.error(request, 'Quote not found')
        
        else:
            # Upload new quote
            image = request.FILES.get('image')
            quote_text = request.POST.get('quote', '').strip()
            author = request.POST.get('author', '').strip() or 'Unknown'
            
            if not image:
                messages.error(request, 'Image is required!')
            elif image.size > 5 * 1024 * 1024:
                messages.error(request, 'Image size must be less than 5MB')
            else:
                MotivationalQuote.objects.create(
                    quote=quote_text if quote_text else None,
                    author=author,
                    image=image
                )
                messages.success(request, 'Quote uploaded successfully!')
        
        return redirect('aquaguard:admin-quotes-management')
    
    # GET request - show all quotes
    quotes = MotivationalQuote.objects.all().order_by('-created_at')
    
    context = {
        'quotes': quotes,
        'total_quotes': quotes.count(),
        'active_quotes': quotes.filter(is_active=True).count(),
        'inactive_quotes': quotes.filter(is_active=False).count(),
    }
    return render(request, 'aquaguard/admin_quotes_management.html', context)



#================Profile Edit Requests Management Views============================="""

@login_required

def admin_edit_requests(request):
    """Admin view to see all profile edit requests"""
    from .models import ProfileEditRequest
    
    status_filter = request.GET.get('status', 'pending')
    if status_filter == 'all':
        edit_requests = ProfileEditRequest.objects.all()
    else:
        edit_requests = ProfileEditRequest.objects.filter(status=status_filter)
    
    context = {
        'edit_requests': edit_requests,
        'status_filter': status_filter,
    }
    return render(request, 'aquaguard/admin_edit_requests.html', context)

@login_required
def admin_process_edit_request(request, request_id):
    """Admin processes a profile edit request"""
    from .models import ProfileEditRequest
    from django.utils import timezone
    
    edit_request = ProfileEditRequest.objects.get(id=request_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            edit_request.status = 'approved'
            edit_request.processed_at = timezone.now()
            edit_request.processed_by = request.user.email
            edit_request.admin_notes = request.POST.get('admin_notes', '')
            edit_request.save()
            messages.success(request, f"Edit request for {edit_request.customer.name} approved. You can now edit their profile.")
            return redirect('aquaguard:admin-edit-customer-profile', customer_id=edit_request.customer.id, request_id=request_id)
        
        elif action == 'reject':
            edit_request.status = 'rejected'
            edit_request.processed_at = timezone.now()
            edit_request.processed_by = request.user.email
            edit_request.admin_notes = request.POST.get('admin_notes', '')
            edit_request.save()
            messages.info(request, f"Edit request for {edit_request.customer.name} rejected.")
            return redirect('aquaguard:admin-edit-requests')
    
    context = {
        'edit_request': edit_request,
    }
    return render(request, 'aquaguard/admin_process_edit_request.html', context)


#================Admin Edit Customer Profile View function============================="""

@login_required
def admin_edit_customer_profile(request, customer_id, request_id):
    """Admin edits customer profile after approving request"""
    from .models import ProfileEditRequest
    
    customer = Customer.objects.get(id=customer_id)
    edit_request = None
    if int(request_id) != 0:
        edit_request = ProfileEditRequest.objects.get(id=request_id)

    if request.method == 'POST':
        # Update only allowed fields and keep required ones intact if left blank
        phone_number = request.POST.get('phone_number', '').strip()
        customer.phone_number = phone_number or customer.phone_number

        token_input = request.POST.get('device_token')
        if token_input is None:
            token_input = request.POST.get('token')
        if token_input is not None:
            new_token = token_input.strip()
            if new_token:
                customer.device_token = new_token

        # Keep existing device chip ID unless explicitly provided (request form hides it)
        device_chip_id = request.POST.get('device_chip_id')
        if device_chip_id:
            device_chip_id = device_chip_id.strip()
            if not device_chip_id:
                messages.error(request, "Device chip ID cannot be empty.")
                return redirect(request.path)
            customer.device_chip_id = device_chip_id

        firmware_version_input = request.POST.get('firmware_version')
        if firmware_version_input is None:
            firmware_version_input = request.POST.get('ota_version')
        if firmware_version_input is not None:
            firmware_value = firmware_version_input.strip()
            customer.firmware_version = firmware_value or None
            if hasattr(customer, 'ota_version'):
                customer.ota_version = firmware_value or None

        family_members = request.POST.get('number_of_family_members')
        if family_members:
            try:
                customer.number_of_family_members = int(family_members)
            except ValueError:
                messages.error(request, "Family members must be a number.")
                return redirect(request.path)

        customer.customer_address = request.POST.get('customer_address', customer.customer_address)

        paisa_per_litre = request.POST.get('paisa_per_litre')
        if paisa_per_litre:
            try:
                customer.paisa_per_litre = float(paisa_per_litre)
            except ValueError:
                messages.error(request, "Invalid paisa per litre value.")
                return redirect(request.path)

        board_value = request.POST.get('board')
        if board_value:
            if board_value not in BOARD_OPTION_VALUES:
                messages.error(request, "Invalid board selection.")
                return redirect(request.path)
            customer.board = board_value

        customer_type = request.POST.get('customer_type')
        if customer_type:
            valid_types = dict(Customer.CUSTOMER_TYPE_CHOICES)
            if customer_type not in valid_types:
                messages.error(request, "Invalid customer type selected.")
                return redirect(request.path)
            customer.customer_type = customer_type

        customer.device_token = request.POST.get('device_token', '').strip() or None
        customer.device_health = request.POST.get('device_health', '').strip() or None

        customer.save()
        messages.success(request, f"Profile updated successfully for {customer.name}")
        return redirect('aquaguard:customer-details', customer_id=customer.id)

    context = {
        'customer': customer,
        'edit_request': edit_request,
    }
    return render(request, 'aquaguard/admin_edit_customer_profile.html', context)

#================Sensor Data Ingestion API View function============================="""
@csrf_exempt
@api_view(['POST'])
def ingest_sensor_data(request):
    raw_data = request.data

    # -------- Detect mode --------
    if "data" in raw_data:
        serializer = SensorBatchSerializer(data=raw_data)
        mode = "batch"
    else:
        serializer = SensorDataSerializer(data=raw_data)
        mode = "single"

    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    payload = serializer.validated_data
    device_id = payload["device_id"]
    ts = payload["ts"]
    signature = payload["signature"]

    # -------- Replay protection --------
    if abs(time.time() - ts) > 60:
        return Response({"error": "Request expired"}, status=status.HTTP_403_FORBIDDEN)

    # -------- Get token (cache → DB) --------
    token = get_device_token(device_id)
    if not token:
        return Response({"error": "Invalid device"}, status=status.HTTP_403_FORBIDDEN)

    # -------- Resolve ACTIVE user (CRITICAL FIX) --------
    user = (
        Customer.objects
        .filter(device_chip_id=device_id, is_active=True, block_unblock=True)
        .only("id", "device_token")
        .first()
    )

    if not user:
        return Response({"error": "Invalid device"}, status=status.HTTP_403_FORBIDDEN)

    # -------- Build ORIGINAL payload JSON (for signature) --------
    if mode == "batch":
        original_data_json = json.dumps(
            raw_data["data"],
            separators=(",", ":"),
            ensure_ascii=False,
            sort_keys=False
        )
    else:
        single_data = {
            "water_consumption": raw_data["water_consumption"],
            "tds": raw_data["tds"],
            "ph": raw_data["ph"]
        }
        original_data_json = json.dumps(
            single_data,
            separators=(",", ":"),
            ensure_ascii=False,
            sort_keys=False
        )

    # -------- Verify signature --------
    msg = f"{device_id}{ts}{original_data_json}"
    expected_sig = hmac.new(
        token.encode(),
        msg.encode(),
        hashlib.sha256
    ).hexdigest()

    if not hmac.compare_digest(expected_sig, signature):
        return Response({"error": "Invalid signature"}, status=status.HTTP_403_FORBIDDEN)

    # -------- Update last_seen --------
    #Customer.objects.filter(id=user.id).update(last_seen=django_timezone.now())
    Customer.objects.filter(id=user.id).update(last_seen=django_timezone.now()),


    # -------- Async save (PASS UID, NOT device_id) --------
    if mode == "batch":
        save_sensor_batch.delay(str(user.id), payload["data"])
    else:
        save_sensor_data.delay(
            str(user.id),
            payload["water_consumption"],
            payload["tds"],
            payload["ph"]
        )

    return Response({"status": "accepted"}, status=status.HTTP_202_ACCEPTED)

#================Daily and Monthly Consumption API View functions============================="""
@api_view(["GET"])
def daily_consumption(request):
    """Get daily water consumption for a customer"""
    serializer = DailyConsumptionSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)

    customer_id = serializer.validated_data["uid"]
    date = serializer.validated_data["date"]

    try:
        customer = Customer.objects.get(id=customer_id, is_active=True, block_unblock=True)
    except Customer.DoesNotExist:
        return Response({"error": "Invalid or inactive customer"}, status=status.HTTP_404_NOT_FOUND)

    start = django_timezone.make_aware(datetime.combine(date, datetime.min.time()))
    end = start + timedelta(days=1)

    total = (
        SensorData.objects
        .filter(customer=customer, esp_timestamp__range=(start, end))
        .aggregate(total_liters=Sum("current_volume"))
        .get("total_liters") or 0
    )

    return Response({
        "customer_id": str(customer_id),
        "date": date,
        "daily_consumption_liters": round(float(total), 3)
    })

#================Monthly Consumption API View function============================="""
@api_view(["GET"])
def monthly_consumption(request):
    """Get monthly water consumption for a customer"""
    serializer = MonthlyConsumptionSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)

    customer_id = serializer.validated_data["uid"]
    year = serializer.validated_data["year"]
    month = serializer.validated_data["month"]

    try:
        customer = Customer.objects.get(id=customer_id, is_active=True, block_unblock=True)
    except Customer.DoesNotExist:
        return Response({"error": "Invalid or inactive customer"}, status=status.HTTP_404_NOT_FOUND)

    total = (
        SensorData.objects
        .filter(
            customer=customer,
            esp_timestamp__year=year,
            esp_timestamp__month=month
        )
        .aggregate(total_liters=Sum("current_volume"))
        .get("total_liters") or 0
    )

    return Response({
        "customer_id": str(customer_id),
        "year": year,
        "month": month,
        "monthly_consumption_liters": round(float(total), 3)
    })


# ================= OTA API Views =================

import time
from django.core.cache import cache
from .models import Firmware
from .utils import verify_signature, generate_token
from .cache_utils import get_device_token, get_device_meta, clear_device_cache
import secrets


@csrf_exempt
@api_view(['POST'])
def provision_device(request):
    serializer = ProvisionDeviceSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    device_id = serializer.validated_data["device_id"]

    # SAFE QUERY - Use Customer model instead of UserInfo
    customer = Customer.objects.filter(
        device_chip_id=device_id,
        is_active=True,
        block_unblock=True
    ).first()

    if not customer:
        return Response(
            {"error": "No active customer assigned to this device"},
            status=404
        )

    if customer.device_token:
        return Response(
            {"error": "Device already provisioned"},
            status=403
        )

    customer.device_token = generate_token()
    customer.save(update_fields=["device_token"])

    # cache by device_chip_id (IMPORTANT)
    cache.set(f"token:{device_id}", customer.device_token, 3600)

    return Response(
        {"token": customer.device_token},
        status=200
    )


@api_view(['GET', 'POST'])
def check_ota(request):
    # Handle both GET (query_params) and POST (data)
    if request.method == 'POST':
        serializer = CheckOTASerializer(data=request.data)
    else:
        serializer = CheckOTASerializer(data=request.query_params)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    device_id = data['device_id']
    version = data['version']
    ts = data['ts']  # Now an integer from serializer validation
    signature = data['signature']

    # Replay protection (±60s)
    if abs(time.time() - ts) > 60:
        return Response(
            {"error": "Request expired"},
            status=status.HTTP_403_FORBIDDEN
        )

    # TOKEN (Redis → DB fallback)
    token = get_device_token(device_id)
    if not token:
        return Response(
            {"error": "Invalid device"},
            status=status.HTTP_403_FORBIDDEN
        )

    # Signature verification
    if not verify_signature(device_id, version, ts, signature, token):
        return Response(
            {"error": "Invalid signature"},
            status=status.HTTP_403_FORBIDDEN
        )

    # DEVICE META (Redis → DB fallback)
    meta = get_device_meta(device_id)
    if not meta:
        return Response({"update": False})

    target_version = meta["firmware_version"]
    board = meta["board"]

    # OTA decision
    if version != target_version:
        fw = Firmware.objects.filter(
            board=board,
            version=target_version,
            is_active=True
        ).only("version", "bin_file").first()

        if not fw:
            return Response({"update": False})

        return Response({
            "update": True,
            "version": fw.version,
            "url": request.build_absolute_uri(fw.bin_file.url)
        })

    return Response({"update": False})


# ========== FORGOT PASSWORD APIs ==========



@api_view(['POST'])
@csrf_exempt
def forgot_password_request(request):
    """
    Request password reset - sends TOTP code to registered email
    POST: Send password reset code
    Required: email
    """
    serializer = ForgotPasswordRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({
            "success": False,
            "message": "Validation failed",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    email = serializer.validated_data['email']
    
    # Check if email exists
    try:
        customer = Customer.objects.get(email=email)
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "No account found with this email address"
        }, status=status.HTTP_404_NOT_FOUND)
    
    try:
        # Generate TOTP secret for password reset
        secret = generate_totp_secret()
        
        # Store secret temporarily for verification
        '''
        _pending_password_resets[email] = {
            'secret': secret,
            'created_at': timezone.now(),
            'customer_id': str(customer.id),
            'verified': False
        }
        
        cache.set(
            f"pwd_reset_{email}",
            {
                "secret": secret,
                "customer_id": str(customer.id),
                "verified": False,
                "created_at": timezone.now().timestamp()
            },
            timeout=300  # 5 minutes
        )
        '''
        PasswordReset.objects.update_or_create(
            email=email,
            defaults={
                "secret": secret,
                "customer": customer,
                "verified": False,
                "attempts": 0,
                "verified_at": None
            }
        )
        
        # Get current TOTP code
        totp_code = get_totp_code(secret)
        
        # Send password reset email
        email_sent = send_password_reset_email(email, totp_code, customer.name)
        
        if email_sent:
            return Response({
                "success": True,
                "message": f"Password reset code sent to {email}",
                "data": {
                    "email": email,
                    "expires_in": 300
                }
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "success": False,
                "message": "Failed to send password reset email. Please try again."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        logger.error(f"Forgot password request error: {str(e)}")
        return Response({
            "success": False,
            "message": f"Failed to process request: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def forgot_password_verify(request):
    """
    Verify TOTP code for password reset
    POST: Verify password reset code
    Required: email, code
    """
    serializer = ForgotPasswordVerifySerializer(data=request.data)
    if not serializer.is_valid():
        return Response({
            "success": False,
            "message": "Validation failed",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    email = serializer.validated_data['email']
    code = serializer.validated_data['code']
    
    # Check if we have a pending reset for this email
    #pending = _pending_password_resets.get(email)
    #pending = cache.get(f"pwd_reset_{email}")
    reset = PasswordReset.objects.filter(email=email).first()

    if not reset:
        return Response({
            "success": False,
            "message": "No pending password reset found. Please request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if reset request is too old (5 minutes max)
    '''
    created_at = pending.get('created_at')
    if created_at and (timezone.now() - created_at).total_seconds() > 300:
        del _pending_password_resets[email]
        return Response({
            "success": False,
            "message": "Password reset request expired. Please request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not reset:
        return Response({
            "success": False,
            "message": "No pending password reset found."
        }, status=400)
    '''

    if reset.is_expired():
        reset.delete()
        return Response({
            "success": False,
            "message": "Password reset request expired."
        }, status=400)
    
    # Verify the TOTP code
    '''
    secret = pending.get('secret')
    if verify_totp_code(secret, str(code)):
        # Mark as verified but don't delete yet - needed for reset step
        #_pending_password_resets[email]['verified'] = True
        pending["verified"] = True
        cache.set(f"pwd_reset_{email}", pending, timeout=600)
        _pending_password_resets[email]['verified_at'] = timezone.now()
        return Response({
            "success": True,
            "message": "Code verified successfully. You can now reset your password.",
            "data": {
                "email": email,
                "verified": True
            }
        }, status=status.HTTP_200_OK)
    '''
    
    if reset.attempts >= 5:
        reset.delete()
        return Response({
            "success": False,
            "message": "Too many attempts. Please request again."
        }, status=400)

    reset.attempts += 1
    reset.save(update_fields=["attempts"])

    if verify_totp_code(reset.secret, str(code)):
        reset.verified = True
        reset.verified_at = timezone.now()
        reset.save(update_fields=["verified", "verified_at"])

        return Response({
            "success": True,
            "message": "Code verified successfully."
        }, status=200)
    else:
        return Response({
            "success": False,
            "message": "Invalid or expired verification code. Please try again or request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@csrf_exempt
def forgot_password_reset(request):
    """
    Reset password after verification
    POST: Reset password
    Required: email, new_password
    """
    serializer = ForgotPasswordResetSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({
            "success": False,
            "message": "Validation failed",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    email = serializer.validated_data['email']
    new_password = serializer.validated_data['new_password']
    
    # Check if we have a verified reset for this email
    #pending = _pending_password_resets.get(email)
    '''
    pending = cache.get(f"pwd_reset_{email}")
    if not pending:
        return Response({
            "success": False,
            "message": "No pending password reset found. Please start the process again."
        }, status=status.HTTP_400_BAD_REQUEST)
    '''
    reset = PasswordReset.objects.filter(email=email).first()
    if not reset or not reset.verified:
        return Response({
            "success": False,
            "message": "Please verify your code first."
        }, status=400)
    '''
    if not pending.get('verified'):
        return Response({
            "success": False,
            "message": "Please verify your code first before resetting password."
        }, status=status.HTTP_400_BAD_REQUEST)
    '''
    
    # Check if verification is too old (10 minutes max from verification)
    '''
    verified_at = pending.get('verified_at')
    if verified_at and (timezone.now() - verified_at).total_seconds() > 600:
        del _pending_password_resets[email]
        return Response({
            "success": False,
            "message": "Verification expired. Please start the process again."
        }, status=status.HTTP_400_BAD_REQUEST)
    '''
    if (timezone.now() - reset.verified_at).total_seconds() > 600:
        reset.delete()
        return Response({
            "success": False,
            "message": "Verification expired. Please start again."
        }, status=400)
    
    try:
        from django.contrib.auth.hashers import make_password
        
        customer = Customer.objects.get(email=email)
        customer.password = make_password(new_password)
        customer.save(update_fields=['password'])
        
        # Clean up
        #del _pending_password_resets[email]
        #cache.delete(f"pwd_reset_{email}")
        reset.delete()
        
        return Response({
            "success": True,
            "message": "Password reset successfully. You can now login with your new password.",
            "data": {
                "email": email
            }
        }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Password reset error: {str(e)}")
        return Response({
            "success": False,
            "message": f"Failed to reset password: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def send_email_verification(request):
    """
    Send TOTP verification code to email for registration
    POST: Send OTP to email
    Required: email, name (optional for personalized email)
    """
    email = request.data.get('email')
    name = request.data.get('name', '')
    
    if not email:
        return Response({
            "success": False,
            "message": "Email is required"
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if email already exists
    if Customer.objects.filter(email=email).exists():
        return Response({
            "success": False,
            "message": "Email already registered. Please login instead."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Generate TOTP secret for this email
        secret = generate_totp_secret()
        
        # Store secret temporarily (in production use Redis with TTL)
        '''
        _pending_email_verifications[email] = {
            'secret': secret,
            'created_at': timezone.now(),
            'name': name
        }
        '''
        EmailVerification.objects.update_or_create(
            email=email,
            defaults={
                "secret": secret,
                "name": name
            }
        )
        
        # Get current TOTP code
        totp_code = get_totp_code(secret)
        
        # Send email with TOTP code
        email_sent = send_totp_email(email, totp_code, name)
        
        if email_sent:
            return Response({
                "success": True,
                "message": f"Verification code sent to {email}",
                "data": {
                    "email": email,
                    "expires_in": 30
                }
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "success": False,
                "message": "Failed to send verification email. Please try again."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        logger.error(f"Send email verification error: {str(e)}")
        return Response({
            "success": False,
            "message": f"Failed to send verification: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def verify_email_totp(request):
    """
    Verify TOTP code for email verification
    POST: Verify OTP code
    Required: email, code
    """
    email = request.data.get('email')
    code = request.data.get('code')
    
    if not email or not code:
        return Response({
            "success": False,
            "message": "Email and verification code are required"
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if we have a pending verification for this email
    '''
    pending = _pending_email_verifications.get(email)
    if not pending:
        return Response({
            "success": False,
            "message": "No pending verification found. Please request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)
    '''
    verification = EmailVerification.objects.filter(email=email).first()

    if not verification:
        return Response({
            "success": False,
            "message": "No pending verification found. Please request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)
        
    # Check if verification request is too old (5 minutes max)
    #created_at = pending.get('created_at')
    created_at = verification.created_at
    if created_at and (timezone.now() - created_at).total_seconds() > 300:
        del _pending_email_verifications[email]
        return Response({
            "success": False,
            "message": "Verification request expired. Please request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Verify the TOTP code
    #secret = pending.get('secret')
    secret = verification.secret
    if verify_totp_code(secret, str(code)):
        #del _pending_email_verifications[email]
        verification.delete()
        return Response({
            "success": True,
            "message": "Email verified successfully",
            "data": {
                "email": email,
                "verified": True
            }
        }, status=status.HTTP_200_OK)
    else:
        return Response({
            "success": False,
            "message": "Invalid or expired verification code. Please try again or request a new code."
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@csrf_exempt
def resend_email_verification(request):
    """
    Resend TOTP verification code to email
    """
    email = request.data.get('email')
    
    if not email:
        return Response({
            "success": False,
            "message": "Email is required"
        }, status=status.HTTP_400_BAD_REQUEST)
    
    pending = _pending_email_verifications.get(email, {})
    name = pending.get('name', '')
    
    try:
        secret = generate_totp_secret()
        _pending_email_verifications[email] = {
            'secret': secret,
            'created_at': timezone.now(),
            'name': name
        }
        
        totp_code = get_totp_code(secret)
        email_sent = send_totp_email(email, totp_code, name)
        
        if email_sent:
            return Response({
                "success": True,
                "message": f"New verification code sent to {email}",
                "data": {"email": email, "expires_in": 30}
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                "success": False,
                "message": "Failed to send verification email."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to resend verification: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def customer_registration(request):
    """
    Customer registration API for Android app
    POST: Register new customer
    change 1
    """
    # If a device_chip_id is provided and already assigned to another customer,
    # perform the reassignment step BEFORE running serializer validation so the
    # serializer's unique checks don't block the request.
    from django.db import transaction

    device_chip_id = request.data.get('device_chip_id')
    if device_chip_id:
        existing = Customer.objects.filter(device_chip_id=device_chip_id).first()
        if existing:
            # If the existing owner is active/unblocked, prevent reuse
            if existing.block_unblock:
                return Response({
                    "success": False,
                    "message": "Device already assigned to an active customer"
                }, status=status.HTTP_400_BAD_REQUEST)
            # Otherwise (blocked), rename the old customer's chip id to free it up
            # Use a unique suffix to avoid violating the DB unique constraint.
            with transaction.atomic():
                existing.device_chip_id = f"{existing.device_chip_id}__reassigned__{existing.id}"
                existing.save(update_fields=['device_chip_id'])

    # Ensure any client-sent device token is ignored (device tokens must be
    # provisioned by the ESP during provisioning). Work on a mutable copy
    # because `request.data` can be an immutable QueryDict.
    try:
        incoming = request.data.copy()
    except Exception:
        incoming = dict(request.data)
    # Remove fields that must not be accepted from Android registration
    incoming.pop('device_token', None)
    incoming.pop('token', None)
    serializer = CustomerRegistrationSerializer(data=incoming)

    if serializer.is_valid():
        try:
            email_verified = request.data.get('email_verified', False)

            # Now create the new customer
            customer = serializer.save()

            if email_verified:
                customer.email_verified = True
                customer.save(update_fields=['email_verified'])

            return Response({
                "success": True,
                "message": "Customer registered successfully",
                "data": {
                    "customer_id": str(customer.id),
                    "device_token": customer.device_token,
                    "email_verified": customer.email_verified
                }
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                "success": False,
                "message": f"Registration failed: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # Log validation failure with sanitized input for debugging
    try:
        sanitized = dict(request.data)
        if 'password' in sanitized:
            sanitized['password'] = '********'
    except Exception:
        sanitized = None
    logger.error("Customer registration validation failed. data=%s, errors=%s", sanitized, serializer.errors)

    return Response({
        "success": False,
        "message": "Validation failed",
        "errors": serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)



"""================Customer Registration and Login API View functions============================="""

@api_view(['POST'])
@csrf_exempt
def customer_login(request):
    """
    Customer login API for Android app
    POST: Authenticate customer with email and password
    """
    serializer = CustomerLoginSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response({
            "success": False,
            "message": "Invalid input data",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    email = serializer.validated_data['email']
    password = serializer.validated_data['password']
    
    try:
        # Check if customer exists
        customer = Customer.objects.get(email=email)
        
        # Check password
        from django.contrib.auth.hashers import check_password
        if not check_password(password, customer.password):
            return Response({
                "success": False,
                "message": "Password does not match"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Successful login - return customer profile
        profile_serializer = CustomerProfileSerializer(customer)
        return Response({
            "success": True,
            "message": "Login successful",
            "data": profile_serializer.data
        }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Email does not exist"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Login failed: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


# ================= Customer Profile Update API ==================

@api_view(['PUT', 'PATCH'])
@csrf_exempt
def customer_profile_update(request, customer_id):
    """
    Customer profile update API for Android app

    PUT / PATCH: Update customer profile

    Required:
        - customer_id (UUID)

    Optional fields:
        - name
        - phone_number
        - customer_address
        - number_of_family_members
        - location
        - latitude
        - longitude
    """

    # Fetch customer
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Customer not found"
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # Check if customer account is blocked
    if not customer.block_unblock:
        return Response(
            {
                "success": False,
                "message": "Account is blocked. Please contact support."
            },
            status=status.HTTP_403_FORBIDDEN
        )

    # Initialize serializer
    serializer = CustomerProfileUpdateSerializer(
        customer,
        data=request.data,
        partial=True
    )

    # Validate and save
    if serializer.is_valid():
        try:
            serializer.save()

            # Return updated profile
            profile_serializer = CustomerProfileSerializer(customer)

            return Response(
                {
                    "success": True,
                    "message": "Profile updated successfully",
                    "data": profile_serializer.data
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Failed to update profile: {str(e)}"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    return Response(
        {
            "success": False,
            "message": "Validation failed",
            "errors": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )


#================Customer Info API View function============================="""
@api_view(['GET'])
@csrf_exempt
def customer_info(request, customer_id):
    """
    Customer info API for Android app
    GET: Get customer information by UUID
    """
    try:
        # customer_id is already a UUID object from URL parsing (<uuid:customer_id>)
        customer_uuid = customer_id
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Return customer info
        info_serializer = CustomerInfoSerializer(customer)
        return Response({
            "success": True,
            "message": "Customer info retrieved successfully",
            "data": info_serializer.data
        }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to retrieve customer info: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#================Customer Location Update API View function============================="""
@api_view(['GET', 'POST'])
@csrf_exempt
def customer_location(request, customer_id):
    """
    Customer location API for Android app
    GET: Retrieve customer location data by UUID
    POST: Update customer location data by UUID
    """
    try:
        # customer_id is already a UUID object from URL parsing
        customer_uuid = customer_id
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        if request.method == 'GET':
            # Return current location data
            return Response({
                "success": True,
                "message": "Location retrieved successfully",
                "data": {
                    "latitude": customer.latitude,
                    "longitude": customer.longitude,
                    "location": customer.location
                }
            }, status=status.HTTP_200_OK)
        
        elif request.method == 'POST':
            # Validate location data
            serializer = CustomerLocationSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "success": False,
                    "message": "Invalid location data",
                    "errors": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Update customer location
            customer.latitude = serializer.validated_data['latitude']
            customer.longitude = serializer.validated_data['longitude']
            customer.location = serializer.validated_data['location']
            customer.save(update_fields=['latitude', 'longitude', 'location'])
            
            return Response({
                "success": True,
                "message": "Location updated successfully",
                "data": {
                    "latitude": customer.latitude,
                    "longitude": customer.longitude,
                    "location": customer.location
                }
            }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to process location request: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#================Customer Water Usage API View function============================="""
'''
@api_view(['GET'])
@csrf_exempt
def customer_water_usage(request, customer_id):
    """
    Customer water usage API for Android app
    GET: Get water usage statistics by customer UUID
    """
    try:
        # customer_id is already a UUID object from URL parsing (<uuid:customer_id>)
        customer_uuid = customer_id
                
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the latest water usage record for this customer
        try:
            #latest_usage = WaterUsage.objects.filter(customer=customer).latest('usage_date')
            latest_usage = WaterUsage.objects.filter(customer=customer).order_by('-usage_date').first()
            
            # Serialize the data
            serializer = CustomerWaterUsageSerializer(latest_usage)
            return Response({
                "success": True,
                "message": "Water usage data retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
            
        except WaterUsage.DoesNotExist:
            # If no usage data exists, return zeros
            return Response({
                "success": True,
                "message": "No water usage data available",
                "data": {
                    "customer": customer.name,
                    "last_24_usage": 0.0,
                    "monthly_usage": 0.0,
                    "per_day_average": 0.0,
                    "current_day_usage": 0.0
                }
            }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to retrieve water usage data: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

'''

@api_view(['GET'])
@csrf_exempt
def customer_water_usage(request, customer_id):
    """
    Customer water usage API for Android app
    GET: Get water usage statistics by customer UUID
    """
    try:
        # customer_id is already a UUID object from URL parsing (<uuid:customer_id>)
        customer_uuid = customer_id
                
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get today's date
        today = timezone.now().date()
        today_usage = WaterUsage.objects.filter(customer=customer, usage_date=today).first()
        if today_usage:
            serializer = CustomerWaterUsageSerializer(today_usage)
            return Response({
                "success": True,
                "message": "Water usage data retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        else:
            # Calculate monthly usage and per day average for last 30 days
            from datetime import timedelta
            month_window_start = today - timedelta(days=29)
            month_usages = WaterUsage.objects.filter(customer=customer, usage_date__gte=month_window_start, usage_date__lte=today)
            monthly_total = month_usages.aggregate(total=Sum('current_day_usage'))['total'] or 0.0
            per_day_avg = monthly_total / 30.0
            return Response({
                "success": True,
                "message": "No water usage data available for today",
                "data": {
                    "customer": customer.name,
                    "last_24_usage": 0.0,
                    "monthly_usage": monthly_total,
                    "per_day_average": per_day_avg,
                    "current_day_usage": 0.0
                }
            }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to retrieve water usage data: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#================Customer Recharge Info API View function============================="""
@api_view(['GET'])
@csrf_exempt
def customer_recharge_info(request, customer_id):
    """
    Customer recharge info API for Android app
    GET: Get recharge information by customer UUID
    """
    try:
        # customer_id is already a UUID object from URL parsing (<uuid:customer_id>)
        customer_uuid = customer_id
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the latest/active recharge record for this customer
        try:
            # First try to get active recharge, if none, get the latest one
            '''
            recharge = Recharge.objects.filter(
                customer=customer, 
                status='active'
            ).first()
            
            if not recharge:
                # If no active recharge, get the latest one
                recharge = Recharge.objects.filter(customer=customer).latest('recharge_date')
            
            # Serialize the data
            serializer = CustomerRechargeInfoSerializer(recharge)
            '''
            
            recharges = Recharge.objects.filter(customer=customer).order_by('-recharge_date')

            if not recharges.exists():
                return Response({
                    "success": False,
                    "message": "No recharge information available"
                }, status=status.HTTP_404_NOT_FOUND)

            serializer = CustomerRechargeInfoSerializer(recharges, many=True)
            return Response({
                "success": True,
                "message": "Recharge info retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
            
        except Recharge.DoesNotExist:
            # If no recharge data exists
            return Response({
                "success": False,
                "message": "No recharge information available"
            }, status=status.HTTP_404_NOT_FOUND)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to retrieve recharge info: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#================Customer Payment Details API View function============================="""
@api_view(['GET', 'POST'])
@csrf_exempt
def customer_payment_details(request, customer_id):
    """
    Customer payment details API for Android app
    GET: Retrieve customer payment history by UUID
    POST: Store payment information by customer UUID
    """
    try:
        # customer_id is already a UUID object from URL parsing
        customer_uuid = customer_id
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        if request.method == 'GET':
            # Retrieve customer's payment history
            payments = Payment.objects.filter(customer=customer).order_by('-recharge_date')
            serializer = PaymentSerializer(payments, many=True)
            return Response({
                "success": True,
                "message": "Payment details retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        
        elif request.method == 'POST':
            # Validate payment data
            serializer = PaymentCreateSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "success": False,
                    "message": "Invalid payment data",
                    "errors": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if order_id already exists
            order_id = serializer.validated_data['order_id']
            if Payment.objects.filter(order_id=order_id).exists():
                return Response({
                    "success": False,
                    "message": "Order ID already exists"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create payment record
            payment = Payment.objects.create(
                customer=customer,
                amount=serializer.validated_data['amount'],
                order_id=order_id,
                payment_status=serializer.validated_data['payment_status'],
                transaction_id=serializer.validated_data.get('transaction_id', ''),
            )
            
            # If recharge_date is provided, update it
            if 'recharge_date' in serializer.validated_data:
                payment.recharge_date = serializer.validated_data['recharge_date']
                payment.save()
            
            # If customer has a subscription plan and payment is completed, create a recharge record
            if customer.subscription_plan_id and payment.payment_status == 'completed':
                from datetime import timedelta
                from django.utils import timezone
                
                # Create recharge record
                recharge = Recharge.objects.create(
                    customer=customer,
                    litres_allocated=customer.subscription_plan.litres_allocated,
                    litres_used=0,
                    litres_remaining=customer.subscription_plan.litres_allocated,
                    expiry_date=timezone.now() + timedelta(days=customer.subscription_plan.duration_days),
                    status='active'
                )
                
                # Update customer's last_recharge_date
                customer.last_recharge_date = payment.recharge_date
                customer.save(update_fields=['last_recharge_date'])
            
            return Response({
                "success": True,
                "message": "Payment details stored successfully",
                "data": {
                    "payment_id": payment.id,
                    "customer": customer.name,
                    "email": customer.email,
                    "amount": str(payment.amount),
                    "recharge_date": payment.recharge_date.isoformat(),
                    "order_id": payment.order_id,
                    "payment_status": payment.payment_status,
                    "transaction_id": payment.transaction_id
                }
            }, status=status.HTTP_201_CREATED)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to process payment request: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#================Subscription Info API View function============================="""
@api_view(['GET'])
@csrf_exempt
def subscription_info(request):
    """
    Subscription info API for Android app
    GET: Get all available subscription plans (public API, no authentication required)
    """
    try:
        # Get all active subscription plans
        subscription_plans = SubscriptionPlan.objects.filter(status=True)
        
        # Serialize the data
        serializer = SubscriptionInfoSerializer(subscription_plans, many=True)
        
        return Response({
            "success": True,
            "message": "Subscription plans retrieved successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to retrieve subscription plans: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
#================Customer Subscription API View function============================="""
@api_view(['POST'])
@csrf_exempt
def customer_subscribe_plan(request, customer_id):
    """
    Customer subscribe to plan API for Android app
    POST: Subscribe customer to a subscription plan by UUID
    Creates Payment and Recharge records
    Note: Blocked customers CAN subscribe - this unblocks them upon successful payment
    """
    try:
        import uuid
        from django.db import transaction as db_transaction
        
        # customer_id is already a UUID object from URL parsing
        customer_uuid = customer_id
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Note: We allow blocked customers to subscribe - successful payment unblocks them
        
        # Check if customer type allows subscription
        if customer.customer_type != 'subscription':
            return Response({
                "success": False,
                "message": "Only subscription customers can subscribe to plans."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate subscription plan data
        plan_id = request.data.get('subscription_plan_id')
        payment_method = request.data.get('payment_method', 'online')
        transaction_id = request.data.get('transaction_id', '')
        # Accept order_id from POST data (for Android payment)
        incoming_order_id = request.data.get('order_id')
        
        if not plan_id:
            return Response({
                "success": False,
                "message": "subscription_plan_id is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            subscription_plan = SubscriptionPlan.objects.get(id=plan_id, status=True)
        except SubscriptionPlan.DoesNotExist:
            return Response({
                "success": False,
                "message": "Invalid or inactive subscription plan"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Use database transaction to ensure atomicity
        with db_transaction.atomic():
            # Subscribe customer to the plan
            customer.subscription_plan = subscription_plan
            customer.block_unblock = True  # Ensure customer is unblocked
            customer.save(update_fields=['subscription_plan', 'block_unblock'])
            
            # Use order_id from POST data if provided, else generate unique order ID
            order_id = incoming_order_id or f"ORD-{uuid.uuid4().hex[:12].upper()}"
            
            # Create Payment record
            payment = Payment.objects.create(
                customer=customer,
                amount=subscription_plan.price,
                order_id=order_id,
                payment_status='completed',
                transaction_id=transaction_id or f"TXN-{uuid.uuid4().hex[:10].upper()}"
            )
            
            # Calculate expiry date based on plan duration
            expiry_date = timezone.now() + timedelta(days=subscription_plan.duration_days)
            
            # Expire any existing active recharges
            Recharge.objects.filter(customer=customer, status='active').update(status='expired')
            
            # Create Recharge record
            recharge = Recharge.objects.create(
                customer=customer,
                payment=payment, 
                litres_allocated=subscription_plan.litres_allocated,
                litres_used=0,
                litres_remaining=subscription_plan.litres_allocated,
                expiry_date=expiry_date,
                status='active'
            )
            
            # Update customer's last_recharge_date
            customer.last_recharge_date = timezone.now()
            customer.save(update_fields=['last_recharge_date'])
        
        return Response({
            "success": True,
            "message": "Successfully subscribed to plan",
            "data": {
                "customer_id": str(customer.id),
                "subscription_plan": subscription_plan.plan_name,
                "price": str(subscription_plan.price),
                "litres_allocated": subscription_plan.litres_allocated,
                "duration_days": subscription_plan.duration_days,
                "payment": {
                    "order_id": payment.order_id,
                    "amount": str(payment.amount),
                    "status": payment.payment_status,
                    "transaction_id": payment.transaction_id
                },
                "recharge": {
                    "litres_allocated": recharge.litres_allocated,
                    "litres_remaining": recharge.litres_remaining,
                    "expiry_date": recharge.expiry_date.isoformat(),
                    "status": recharge.status
                }
            }
        }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to subscribe to plan: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#================Customer Complaint API View function============================="""
@api_view(['GET', 'POST'])
@csrf_exempt
def customer_complaint(request, customer_id):
    """
    Customer complaint API for Android app
    GET: Retrieve customer complaints by UUID
    POST: Create customer complaint by UUID
    """
    try:
        # Validate UUID format
        customer_uuid = customer_id
        
        # Get customer by UUID
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer is active/blocked
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Account is blocked. Please contact support."
            }, status=status.HTTP_403_FORBIDDEN)
        
        if request.method == 'GET':
            # Retrieve customer's complaints
            complaints = UserComplain.objects.filter(customer=customer).order_by('-complain_date')
            serializer = UserComplainCustomerSerializer(complaints, many=True)
            return Response({
                "success": True,
                "message": "Complaints retrieved successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        
        elif request.method == 'POST':
            # Validate complaint data
            serializer = UserComplainCreateSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "success": False,
                    "message": "Invalid complaint data",
                    "errors": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Generate unique ticket number
            import random
            import string
            ticket_number = f"TKT-{customer.id.hex[:8].upper()}-{random.randint(1000, 9999)}"
            
            # Ensure ticket number is unique
            while UserComplain.objects.filter(ticket_number=ticket_number).exists():
                ticket_number = f"TKT-{customer.id.hex[:8].upper()}-{random.randint(1000, 9999)}"
            
            # Create complaint
            complaint = UserComplain.objects.create(
                customer=customer,
                phone_no=serializer.validated_data['phone_no'],
                problem_statement=serializer.validated_data['problem_statement'],
                priority=serializer.validated_data.get('priority', 'medium'),
                ticket_number=ticket_number,
                future_field1=customer.email,  # Store email in future_field1
            )
            
            return Response({
                "success": True,
                "message": "Complaint submitted successfully",
                "data": {
                    "complaint_id": complaint.id,
                    "customer": customer.name,
                    "phone_no": complaint.phone_no,
                    "email": customer.email,
                    "ticket_number": complaint.ticket_number,
                    "problem_statement": complaint.problem_statement,
                    "complain_date": complaint.complain_date.isoformat(),
                    "priority": complaint.priority,
                    "status": complaint.status
                }
            }, status=status.HTTP_201_CREATED)
        
        else:
            return Response({
                "success": False,
                "message": "Method not allowed"
            }, status=status.HTTP_405_METHOD_NOT_ALLOWED)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to process request: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ========== ESP DEVICE APIs health status post==========
@api_view(['POST'])
@csrf_exempt
def esp_device_health(request):
    """
    ESP Device Health Status API
    POST: ESP device reports its health status
    Verifies device_id, ts, signature, token, then updates device health in DB
    """
    try:
        # Validate request data
        serializer = ESPDeviceHealthSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "success": False,
                "message": "Invalid request data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        device_id = serializer.validated_data['device_id']
        ts = serializer.validated_data['ts']
        signature = serializer.validated_data['signature']
        token = serializer.validated_data['token']
        health_status = serializer.validated_data['health_status']

        # Verify signature: device_id + ts + health_status signed with token
        ts_str = str(int(ts))
        msg = f"{device_id}{ts_str}{health_status}".encode()
        expected_sig = hmac.new(
            token.encode(),
            msg,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, signature):
            return Response({
                "success": False,
                "message": "Invalid signature"
            }, status=status.HTTP_403_FORBIDDEN)

        # Get customer by device_id and verify token matches
        try:
            customer = Customer.objects.get(
                device_chip_id=device_id,
                device_token=token,
                is_active=True
            )
        except Customer.DoesNotExist:
            return Response({
                "success": False,
                "message": "Device not found or invalid token"
            }, status=status.HTTP_404_NOT_FOUND)

        # Update device health status and last_seen timestamp
        Customer.objects.filter(id=customer.id).update(
            device_health=health_status,
            last_seen=django_timezone.now()
        )

        # Return success response
        return Response({
            "success": True,
            "device_id": device_id,
            "health_status": health_status,
            "timestamp": int(time.time()),
            "message": "Device health status updated successfully"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to update device health: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ========== ESP DEVICE APIs calibration factor get==========
@api_view(['GET', 'POST'])
@csrf_exempt
def esp_calibration_factor(request):
    """
    ESP Device Calibration Factor API
    GET/POST: ESP device requests its calibration factor
    Verifies device_id, ts, signature, token, then returns calibration factor
    """
    try:
        # Handle both GET and POST
        data = request.data if request.method == 'POST' else request.query_params
        serializer = ESPCalibrationFactorSerializer(data=data)
        if not serializer.is_valid():
            return Response({
                "success": False,
                "message": "Invalid request data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        device_id = serializer.validated_data['device_id']
        ts = serializer.validated_data['ts']
        signature = serializer.validated_data['signature']
        token = serializer.validated_data['token']

        # Verify signature: device_id + ts signed with token
        ts_str = str(int(ts))
        msg = f"{device_id}{ts_str}".encode()
        expected_sig = hmac.new(
            token.encode(),
            msg,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, signature):
            return Response({
                "success": False,
                "message": "Invalid signature"
            }, status=status.HTTP_403_FORBIDDEN)

        # Get customer by device_id and verify token matches
        try:
            customer = Customer.objects.get(
                device_chip_id=device_id,
                device_token=token,
                is_active=True
            )
        except Customer.DoesNotExist:
            return Response({
                "success": False,
                "message": "Device not found or invalid token"
            }, status=status.HTTP_404_NOT_FOUND)

        # Get the latest calibration factor for this customer
        try:
            latest_calibration = CalibrationData.objects.filter(
                customer=customer
            ).order_by('-recorded_at').first()

            if latest_calibration:
                calibration_factor = latest_calibration.calibration_factor
                recorded_at = latest_calibration.recorded_at.isoformat()
            else:
                # No calibration data found, return default
                calibration_factor = 1.0  # Default calibration factor
                recorded_at = None

        except Exception as e:
            # If there's an error getting calibration data, return default
            calibration_factor = 1.0
            recorded_at = None

        # Update last_seen timestamp
        Customer.objects.filter(id=customer.id).update(last_seen=django_timezone.now())

        # Return calibration factor response
        return Response({
            "success": True,
            "device_id": device_id,
            "calibration_factor": calibration_factor,
            "recorded_at": recorded_at,
            "timestamp": int(time.time()),
            "message": "Calibration factor retrieved successfully"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to retrieve calibration factor: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def fcm_token_register(request):
    """
    FCM Token Registration API
    POST: Mobile app registers/updates FCM token for push notifications
    """
    try:
        # Validate request data
        serializer = FCMTokenSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "success": False,
                "message": "Invalid request data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        customer_id = serializer.validated_data['customer_id']
        fcm_token = serializer.validated_data['fcm_token']

        # Get customer by UUID
        try:
            customer = Customer.objects.get(id=customer_id, is_active=True)
        except Customer.DoesNotExist:
            return Response({
                "success": False,
                "message": "Customer not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Update FCM token
        Customer.objects.filter(id=customer.id).update(
            fcm_token=fcm_token,
            updated_at=django_timezone.now()
        )

        # Return success response
        return Response({
            "success": True,
            "customer_id": str(customer_id),
            "message": "FCM token registered successfully"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to register FCM token: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ========== ESP DEVICE APIs sensor data post==========
@api_view(['POST'])
@csrf_exempt
def esp_sensor_data(request):
    """
    ESP Device Sensor Data API
    POST: ESP device submits sensor data for storage
    Verifies device_id, ts, signature, token & JSON data, then inserts to DB
    """
    try:
        # Validate request data
        serializer = ESPSensorDataSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "success": False,
                "message": "Invalid request data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        device_id = serializer.validated_data['device_id']
        ts = serializer.validated_data['ts']
        signature = serializer.validated_data['signature']
        token = serializer.validated_data['token']
        # Keep validated data for checks but pass the original JSON payload
        # to Celery to avoid non-serializable Python `datetime` objects.
        sensor_data = serializer.validated_data['data']
        raw_data = request.data.get('data', [])
        sensor_payload = raw_data

        # Verify signature: device_id + ts + JSON data signed with token
        # Create the JSON string exactly as the ESP would send it
        raw_data = request.data.get('data', [])
        data_json = json.dumps(
            raw_data,
            separators=(",", ":"),
            sort_keys=False,
            ensure_ascii=False
        )
        ts_str = str(int(ts))
        msg = f"{device_id}{ts_str}{data_json}".encode()

        expected_sig = hmac.new(
            token.encode(),
            msg,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, signature):
            return Response({
                "success": False,
                "message": "Invalid signature"
            }, status=status.HTTP_403_FORBIDDEN)

        # Get customer by device_id and verify token matches
        try:
            customer = Customer.objects.get(
                device_chip_id=device_id,
                device_token=token,
                is_active=True,
                block_unblock=True  # Only allow active/unblocked devices to submit data
            )
        except Customer.DoesNotExist:
            return Response({
                "success": False,
                "message": "Device not found, invalid token, or device is blocked"
            }, status=status.HTTP_404_NOT_FOUND)

        # Update last_seen timestamp
        Customer.objects.filter(id=customer.id).update(last_seen=django_timezone.now())

        # Async save sensor data (PASS customer ID, NOT device_id)
        # Use the original JSON payload (`sensor_payload`) so Celery receives
        # only JSON-serializable types (ints/strings), not datetimes.
        save_sensor_batch.delay(str(customer.id), sensor_payload)

        # Return success response
        return Response({
            "success": True,
            "device_id": device_id,
            "data_points": len(sensor_payload),
            "timestamp": int(time.time()),
            "message": "Sensor data accepted for processing"
        }, status=status.HTTP_202_ACCEPTED)

    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to process sensor data: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ========== ESP DEVICE APIs ota check ==========
@api_view(['GET', 'POST'])
@csrf_exempt
def esp_ota_status_check(request):
    """
    ESP Device OTA Status Check API
    GET/POST: ESP device requests OTA status information
    Verifies device_id, ts, signature & token
    """
    try:
        # Handle both GET and POST
        data = request.data if request.method == 'POST' else request.query_params
        serializer = ESPOTAStatusCheckSerializer(data=data)
        if not serializer.is_valid():
            return Response({
                "success": False,
                "message": "Invalid request data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        device_id = serializer.validated_data['device_id']
        ts = serializer.validated_data['ts']
        signature = serializer.validated_data['signature']
        token = serializer.validated_data['token']

        # Verify signature: device_id + ts signed with token
        ts_str = str(int(ts))
        msg = f"{device_id}{ts_str}".encode()
        expected_sig = hmac.new(
            token.encode(),
            msg,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, signature):
            return Response({
                "success": False,
                "message": "Invalid signature"
            }, status=status.HTTP_403_FORBIDDEN)

        # Get customer by device_id and verify token matches
        try:
            customer = Customer.objects.get(
                device_chip_id=device_id,
                device_token=token,
                is_active=True
            )
        except Customer.DoesNotExist:
            return Response({
                "success": False,
                "message": "Device not found or invalid token"
            }, status=status.HTTP_404_NOT_FOUND)

        # Check if device is blocked - blocked devices shouldn't get OTA updates
        if not customer.block_unblock:
            return Response({
                "success": False,
                "message": "Device is blocked",
                "ota_available": False
            }, status=status.HTTP_403_FORBIDDEN)

        # Return OTA status information
        # For now, return basic status - can be extended with actual firmware version checking
        return Response({
            "success": True,
            "device_id": device_id,
            "ota_available": True,  # Placeholder - would check against current firmware version
            "current_version": customer.firmware_version or "unknown",
            "board_type": customer.board or "esp32",
            "last_check": int(time.time()),
            "message": "OTA status check successful"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to process OTA status check: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ========== ESP DEVICE APIs block unblock ==========
'''
@api_view(['POST'])
@csrf_exempt
def esp_block_unblock_status(request):
    """
    ESP Device Block/Unblock Status API
    POST: ESP device requests its block/unblock status
    Verifies device_id, ts, signature & token
    """
    try:
        # Validate request data
        serializer = ESPBlockUnblockStatusSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "success": False,
                "message": "Invalid request data",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        device_id = serializer.validated_data['device_id']
        ts = serializer.validated_data['ts']
        signature = serializer.validated_data['signature']
        token = serializer.validated_data['token']

        # Verify signature: device_id + ts signed with token
        ts_str = str(int(ts))
        msg = f"{device_id}{ts_str}".encode()
        expected_sig = hmac.new(
            token.encode(),
            msg,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, signature):
            return Response({
                "success": False,
                "message": "Invalid signature"
            }, status=status.HTTP_403_FORBIDDEN)

        # Get customer by device_id and verify token matches
        try:
            customer = Customer.objects.get(
                device_chip_id=device_id,
                device_token=token,
                is_active=True
            )
        except Customer.DoesNotExist:
            return Response({
                "success": False,
                "message": "Device not found or invalid token"
            }, status=status.HTTP_404_NOT_FOUND)

                # Check if recharge is expired or exhausted - AUTO BLOCK
        # Per-litre customers: only check litres_remaining
        # Subscription customers: check litres AND expiry_date
        if customer.customer_type == 'per_litre':
            has_valid_recharge = Recharge.objects.filter(
                customer=customer,
                status='active',
                litres_remaining__gt=0
            ).exists()
        else:
            has_valid_recharge = Recharge.objects.filter(
                customer=customer,
                status='active',
                litres_remaining__gt=0,
                expiry_date__gt=timezone.now()
            ).exists()

        # Auto block if no valid recharge found
        if not has_valid_recharge:
            customer.block_unblock = False
            customer.save(update_fields=['block_unblock'])

        # Return block/unblock status
        return Response({
            "success": True,
            "device_id": device_id,
            "block_unblock": customer.block_unblock,  # True=Active/Unblocked, False=Blocked/Inactive
            "has_valid_recharge": has_valid_recharge,
            "timestamp": int(time.time()),
            "message": "Block status retrieved" if customer.block_unblock else "Device blocked - recharge expired or exhausted"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to process request: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
'''    
import time
import hmac
import hashlib

from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

# Make sure to import your models and serializers
# from .models import Customer, Recharge
# from .serializers import ESPBlockUnblockStatusSerializer

@api_view(['POST'])
@csrf_exempt
def esp_block_unblock_status(request): 
    """
    ESP Device Block/Unblock Status API
    POST: ESP device requests its block/unblock status. 
    Verifies device_id, ts, signature & token.
    """
    try:
        # Validate request data
        serializer = ESPBlockUnblockStatusSerializer(data=request.data)

        if not serializer.is_valid(): 
            return Response({
                "success": False,
                "message": "Invalid request data", 
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        device_id = serializer.validated_data['device_id'] 
        ts = serializer.validated_data['ts']
        signature = serializer.validated_data['signature'] 
        token = serializer.validated_data['token']

        # Verify signature: device_id + ts signed with token 
        ts_str = str(int(ts))
        msg = f"{device_id}{ts_str}".encode() 
        expected_sig = hmac.new(
            token.encode(), 
            msg, 
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, signature): 
            return Response({
                "success": False,
                "message": "Invalid signature"
            }, status=status.HTTP_403_FORBIDDEN)

        # Get customer by device_id and verify token matches 
        try:
            customer = Customer.objects.get( 
                device_chip_id=device_id, 
                device_token=token, 
                is_active=True
            )
        except Customer.DoesNotExist: 
            return Response({
                "success": False,
                "message": "Device not found or invalid token"
            }, status=status.HTTP_404_NOT_FOUND)

        # Check if recharge is expired or exhausted
        # Per-litre customers: only check litres_remaining
        # Subscription customers: check litres AND expiry_date 
        if customer.customer_type == 'per_litre':
            has_valid_recharge = Recharge.objects.filter( 
                customer=customer,
                status='active', 
                litres_remaining__gt=0  # Fixed: changed 'gt' to '__gt'
            ).exists() 
        else:
            has_valid_recharge = Recharge.objects.filter( 
                customer=customer,
                status='active', 
                litres_remaining__gt=0,         # Fixed: changed 'gt' to '__gt'
                expiry_date__gt=timezone.now()  # Fixed: changed 'gt' to '__gt'
            ).exists()

        # NOTE: We do NOT auto-block customer account when recharge expires. 
        # block_unblock is only set by admin manual action.
        # ESP should use has_valid_recharge to decide if water flow is allowed.

        # Determine if ESP should allow water flow
        # ESP blocked if: admin blocked (block_unblock=False) OR no valid recharge 
        esp_should_block = not customer.block_unblock or not has_valid_recharge
        
        # Determine appropriate status message
        if not customer.block_unblock:
            status_msg = "Admin blocked"
        elif not has_valid_recharge:
            status_msg = "Recharge expired or exhausted"
        else:
            status_msg = "Active"

        # Return status - ESP uses esp_blocked to control water flow 
        return Response({
            "success": True, 
            "device_id": device_id,
            "block_unblock": customer.block_unblock,  # True=Active (admin), False=Blocked by admin
            "has_valid_recharge": has_valid_recharge, # True=Recharge valid, False=Expired/exhausted
            "esp_blocked": esp_should_block,          # True=ESP should stop water, False=Water can flow
            "timestamp": int(time.time()),
            "message": status_msg
        }, status=status.HTTP_200_OK)

    except Exception as e: 
        return Response({
            "success": False,
            "message": f"Failed to process request: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ========== ESP DEVICE APIs ==========

'''
@api_view(['POST'])
def esp_get_user_details(request):
    """
    ESP API: Get user details by device_chip_id

    POST Body:
    {
        "device_chip_id": "..."
    }

    Returns:
        - customer_type
        - expiry_date (if subscription)
        - name
        - email
        - block_unblock
    """

    device_chip_id = request.data.get('device_chip_id')

    if not device_chip_id:
        return Response({
            "success": False,
            "message": "device_chip_id is required"
        }, status=status.HTTP_400_BAD_REQUEST)

    customer = Customer.objects.filter(
        device_chip_id=device_chip_id,
        is_active=True
    ).first()

    if not customer:
        return Response({
            "success": False,
            "message": "Customer not found or inactive"
        }, status=status.HTTP_404_NOT_FOUND)

    # Default: no expiry for per_litre customers
    expiry_date = None
    litres_allocated = 0
    litres_remaining = 0
    litres_used = 0
    
    active_recharge = customer.recharges.filter(
        status='active'
    ).order_by('-recharge_date').first()

    if active_recharge:
        litres_allocated = active_recharge.litres_allocated
        litres_remaining = active_recharge.litres_remaining
        litres_used = active_recharge.litres_used

        if customer.customer_type == 'subscription' and active_recharge.expiry_date:
            expiry_date = active_recharge.expiry_date.isoformat()

    return Response({
        "success": True,
        "customer_type": customer.customer_type,
        "expiry_date": expiry_date,
        "name": customer.name,
        "email": customer.email,
        "block_unblock": customer.block_unblock,
        "litres_allocated": litres_allocated,
        #"litres_remaining": litres_remaining,
        #"litres_used": litres_used
    }, status=status.HTTP_200_OK)
'''
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

# Make sure to import your Customer model here
# from .models import Customer

@api_view(['POST'])
def esp_get_user_details(request):
    """
    ESP API: Get user details by device_chip_id

    POST Body:
    {
        "device_chip_id": "..."
    }

    Returns:
        - customer_type
        - expiry_date (if subscription)
        - name
        - email
        - block_unblock
        - litres_allocated
        - order_id (from latest completed payment)
    """

    device_chip_id = request.data.get('device_chip_id')

    if not device_chip_id:
        return Response({
            "success": False,
            "message": "device_chip_id is required"
        }, status=status.HTTP_400_BAD_REQUEST)

    customer = Customer.objects.filter(
        device_chip_id=device_chip_id,
        is_active=True
    ).first()

    if not customer:
        return Response({
            "success": False,
            "message": "Customer not found or inactive"
        }, status=status.HTTP_404_NOT_FOUND)

    # Default: no expiry for per_litre customers
    expiry_date = None
    litres_allocated = 0
    litres_remaining = 0
    litres_used = 0
    order_id = None
    
    active_recharge = customer.recharges.filter(
        status='active'
    ).order_by('-recharge_date').first()

    if active_recharge:
        litres_allocated = active_recharge.litres_allocated
        litres_remaining = active_recharge.litres_remaining
        litres_used = active_recharge.litres_used

        if customer.customer_type == 'subscription' and active_recharge.expiry_date:
            expiry_date = active_recharge.expiry_date.isoformat()

    # Get the latest completed payment's order_id
    latest_payment = customer.payments.filter(
        payment_status='completed'
    ).order_by('-recharge_date').first()
    
    if latest_payment:
        order_id = latest_payment.order_id

    return Response({
        "success": True,
        "customer_type": customer.customer_type,
        "expiry_date": expiry_date,
        "name": customer.name,
        "email": customer.email,
        "block_unblock": customer.block_unblock,
        "litres_allocated": litres_allocated,
        "order_id": order_id
        # "litres_remaining": litres_remaining,
        # "litres_used": litres_used
    }, status=status.HTTP_200_OK)
    
#================= ESP ONLINE/OFFLINE ===============

import time
import threading

from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

# Make sure to import your Customer model here
# from .models import Customer 

@api_view(['POST'])
@csrf_exempt
def device_ping(request):
    """
    Server pings device and waits for response. 
    POST: { "device_chip_id": "xxxx" }
    """
    device_chip_id = request.data.get("device_chip_id")
    
    if not device_chip_id:
        return Response(
            {"success": False, "message": "device_chip_id required"}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    # Mark device as 'pending'
    Customer.objects.filter(device_chip_id=device_chip_id).update(
        device_status='pending',
        last_seen=timezone.now()
    )

    # Start a background timer to mark offline if no response in 2 minutes 
    def mark_offline():
        time.sleep(120)
        customer = Customer.objects.filter(
            device_chip_id=device_chip_id, 
            device_status='pending'
        ).first()
        
        if customer:
            customer.device_status = 'offline' 
            customer.save(update_fields=['device_status'])
            
    threading.Thread(target=mark_offline, daemon=True).start()

    # Here, you would send a ping to the ESP (e.g., via MQTT, HTTP, etc.) 
    # The ESP should call the below endpoint to confirm online

    return Response(
        {"success": True, "message": "Ping sent, waiting for device response"}, 
        status=status.HTTP_200_OK
    )




from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

# from .models import Customer

@api_view(['POST'])
@csrf_exempt
def device_ping_response(request):
    """
    ESP device responds to ping. 
    POST: { "device_chip_id": "xxxx" } 
    """
    device_chip_id = request.data.get("device_chip_id")
    
    if not device_chip_id:
        return Response(
            {"success": False, "message": "device_chip_id required"}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    customer = Customer.objects.filter(device_chip_id=device_chip_id).first()
    
    if customer:
        customer.device_status = 'online' 
        customer.last_seen = timezone.now()
        customer.save(update_fields=['device_status', 'last_seen'])
        
        return Response(
            {"success": True, "message": "Device marked as online"}, 
            status=status.HTTP_200_OK
        )
    else:
        return Response(
            {"success": False, "message": "Device not found"}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
#================Customer Per-Litre Purchase API View============================="""
@api_view(['POST'])
@csrf_exempt
def customer_purchase_litres(request, customer_id):
    """
    Customer per-litre purchase API for Android app
    POST: Per-litre customer enters amount and gets litres based on their paisa_per_litre rate
    """
    try:
        import uuid
        from django.db import transaction as db_transaction
        
        customer_uuid = customer_id
        customer = Customer.objects.get(id=customer_uuid)
        
        # Check if customer type is per_litre
        if customer.customer_type != 'per_litre':
            return Response({
                "success": False,
                "message": "This API is only for per-litre billing customers. Use subscription API instead."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        amount = request.data.get('amount')
        if not amount:
            return Response({
                "success": False,
                "message": "amount is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError("Amount must be positive")
        except (ValueError, TypeError):
            return Response({
                "success": False,
                "message": "Invalid amount. Must be a positive number."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        paisa_per_litre = customer.paisa_per_litre
        if not paisa_per_litre or paisa_per_litre <= 0:
            return Response({
                "success": False,
                "message": "Paisa per litre rate is not set. Please contact admin."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Calculate litres: amount (rupees) * 100 (to paisa) / paisa_per_litre
        litres_to_add = (amount * 100) / paisa_per_litre
        litres_to_add = round(litres_to_add, 2)
        
        # Per-litre customers: set far future expiry (never expires by date)
        # Only litres_remaining matters for per-litre customers
        expiry_date = timezone.now() + timedelta(days=36500)  # ~100 years (effectively no expiry)
        
        with db_transaction.atomic():
            customer.block_unblock = True
            customer.last_recharge_date = timezone.now()
            customer.save(update_fields=['block_unblock', 'last_recharge_date'])
            
            # Accept order_id from POST data (for Android payment)
            incoming_order_id = request.data.get('order_id')
            order_id = incoming_order_id or f"PPL-{uuid.uuid4().hex[:12].upper()}"
            
            payment = Payment.objects.create(
                customer=customer,
                amount=amount,
                order_id=order_id,
                payment_status='completed',
                transaction_id=request.data.get('transaction_id') or f"TXN-{uuid.uuid4().hex[:10].upper()}"
            )
            
            '''
            active_recharge = Recharge.objects.filter(customer=customer, status='active').first()
            
            if active_recharge:
                active_recharge.litres_allocated += litres_to_add
                active_recharge.litres_remaining += litres_to_add
                if active_recharge.expiry_date < expiry_date:
                    active_recharge.expiry_date = expiry_date
                active_recharge.save()
                recharge = active_recharge
            else:
                recharge = Recharge.objects.create(
                    customer=customer,
                    litres_allocated=litres_to_add,
                    litres_used=0,
                    litres_remaining=litres_to_add,
                    expiry_date=expiry_date,
                    status='active'
                )
            '''
            
            # Get previous active recharge
            active_recharge = Recharge.objects.filter(
                customer=customer,
                status='active'
            ).first()

            previous_remaining = 0

            if active_recharge:
                previous_remaining = active_recharge.litres_remaining
                active_recharge.status = 'expired'
                active_recharge.save(update_fields=['status'])

            # Total litres = previous remaining + new purchase
            total_litres = previous_remaining + litres_to_add

            # Create NEW recharge record
            recharge = Recharge.objects.create(
                customer=customer,
                payment=payment,
                litres_allocated=total_litres,
                litres_used=0,
                litres_remaining=total_litres,
                expiry_date=expiry_date,
                status='active',
                paisa_per_litre_at_purchase=paisa_per_litre,
                litres_purchased=litres_to_add
            )
        
        return Response({
            "success": True,
            "message": f"Successfully purchased {litres_to_add}L for ₹{amount}",
            "data": {
                "customer_id": str(customer.id),
                "amount_paid": amount,
                "paisa_per_litre": paisa_per_litre,
                "litres_purchased": litres_to_add,
                "payment": {
                    "order_id": payment.order_id,
                    "amount": str(payment.amount),
                    "status": payment.payment_status,
                    "transaction_id": payment.transaction_id
                },
                "recharge": {
                    "litres_allocated": recharge.litres_allocated,
                    "litres_remaining": recharge.litres_remaining,
                    "expiry_date": "N/A",  # Per-litre customers don't have expiry
                    "status": recharge.status
                }
            }
        }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            "success": False,
            "message": f"Failed to purchase litres: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    



from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import SensorData, Customer
from .serializers import SensorDataSerializer

#======================== All The Sensor Data ================
from .serializers import SensorDataSerializerNew
from rest_framework.pagination import PageNumberPagination
@api_view(['GET'])
def customer_sensor_data_api(request, customer_id):
    """
    Get all sensor data for a customer (UUID)
    """

    try:
        # Fetch customer
        customer = Customer.objects.get(id=customer_id)

        # Fetch sensor data ordered by timestamp
        sensor_data = (
            SensorData.objects
            .filter(customer=customer)
            .order_by('esp_timestamp')
        )

        # Serialize data
        '''
        serializer = SensorDataSerializerNew(sensor_data, many=True)

        return Response(
            {
                "success": True,
                "count": sensor_data.count(),
                "data": serializer.data,
                "message": "Sensor data fetched successfully"
            },
            status=status.HTTP_200_OK
        )
        '''
        paginator = PageNumberPagination()
        paginator.page_size = 200
        result_page = paginator.paginate_queryset(sensor_data, request)

        serializer = SensorDataSerializerNew(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    except Customer.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Customer not found"
            },
            status=status.HTTP_404_NOT_FOUND
        )

    except Exception as e:
        return Response(
            {
                "success": False,
                "message": f"Failed to fetch sensor data: {str(e)}"
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )



    
    
# ========== CHART DATA APIs FOR ANDROID APP ========== 
 
@api_view(['GET']) 
@csrf_exempt 
def customer_usage_daily_chart(request, customer_id): 
    """ 
    Get daily water usage data for chart visualization 
    GET: Returns chart-ready data for last N days (default 7) 
    Query params: days (1-90, default 7) 
    """ 
    try: 
        customer = Customer.objects.get(id=customer_id) 
         
        if not customer.block_unblock: 
            return Response({ 
                "success": False, 
                "message": "Account is blocked. Please contact support." 
            }, status=status.HTTP_403_FORBIDDEN) 
         
        # Get number of days from query params 
        days = int(request.query_params.get('days', 7)) 
        days = max(1, min(days, 90))  # Clamp between 1 and 90 
         
        today = timezone.now().date() 
        labels = [] 
        values = [] 
         
        # Get usage for each day 
        for i in range(days - 1, -1, -1): 
            target_date = today - timedelta(days=i) 
             
            # Try to get from WaterUsage model first 
            usage = WaterUsage.objects.filter( 
                customer=customer, 
                usage_date=target_date 
            ).first() 
             
            if usage: 
                daily_value = float(usage.current_day_usage or usage.last_24hr_usage or 0) 
            else: 
                # Fallback to SensorData aggregation 
                start = django_timezone.make_aware(datetime.combine(target_date, 
datetime.min.time())) 
                end = start + timedelta(days=1) 
                 
                daily_value = SensorData.objects.filter( 
                    customer=customer, 
                    esp_timestamp__range=(start, end) 
                ).aggregate(total=Sum('current_volume'))['total'] or 0 
                daily_value = float(daily_value) 
             
            # Format label based on number of days 
            if days <= 7: 
                labels.append(target_date.strftime('%a'))  # Mon, Tue, etc. 
            elif days <= 31: 
                labels.append(target_date.strftime('%d %b'))  # 15 Feb 
            else: 
                labels.append(target_date.strftime('%d/%m'))  # 15/02 
             
            values.append(round(daily_value, 2)) 
         
        total = sum(values) 
        average = total / len(values) if values else 0 
        peak_value = max(values) if values else 0 
        peak_index = values.index(peak_value) if peak_value > 0 and values else 0 
        peak_label = labels[peak_index] if labels else '' 
         
        return Response({ 
            "success": True, 
            "message": "Daily usage chart data retrieved successfully", 
            "data": { 
                "labels": labels, 
                "values": values, 
                "total": round(total, 2), 
                "average": round(average, 2), 
                "unit": "litres", 
                "period": f"last_{days}_days", 
                "peak_value": round(peak_value, 2), 
                "peak_label": peak_label 
            } 
        }, status=status.HTTP_200_OK) 
         
    except Customer.DoesNotExist: 
        return Response({ 
            "success": False, 
            "message": "Customer not found" 
        }, status=status.HTTP_404_NOT_FOUND) 
    except Exception as e: 
        logger.error(f"Daily chart data error: {str(e)}") 
        return Response({ 
            "success": False, 
            "message": f"Failed to retrieve chart data: {str(e)}" 
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
 
@api_view(['GET']) 
@csrf_exempt 
def customer_usage_monthly_chart(request, customer_id): 
    """ 
    Get monthly water usage data for chart visualization 
    GET: Returns chart-ready data for last N months (default 12) 
    Query params: months (1-24, default 12) 
    """ 
    try: 
        customer = Customer.objects.get(id=customer_id) 
         
        if not customer.block_unblock: 
            return Response({ 
                "success": False, 
                "message": "Account is blocked. Please contact support." 
            }, status=status.HTTP_403_FORBIDDEN) 
         
        # Get number of months from query params 
        months = int(request.query_params.get('months', 12)) 
        months = max(1, min(months, 24))  # Clamp between 1 and 24 
         
        today = timezone.now().date() 
        labels = [] 
        values = [] 
         
        # Get usage for each month 
        for i in range(months - 1, -1, -1): 
            # Calculate the target month 
            target_date = today - timedelta(days=i * 30) 
            year = target_date.year 
            month = target_date.month 
             
            # Aggregate monthly usage from WaterUsage 
            monthly_usage = WaterUsage.objects.filter( 
                customer=customer, 
                usage_date__year=year, 
                usage_date__month=month 
            ).aggregate(total=Sum('current_day_usage'))['total'] or 0 
             
            if not monthly_usage: 
                # Fallback to SensorData aggregation 
                monthly_usage = SensorData.objects.filter( 
                    customer=customer, 
                    esp_timestamp__year=year, 
                    esp_timestamp__month=month 
                ).aggregate(total=Sum('current_volume'))['total'] or 0 
             
            labels.append(target_date.strftime('%b %Y'))  # Feb 2026 
            values.append(round(float(monthly_usage), 2)) 
         
        total = sum(values) 
        average = total / len(values) if values else 0 
        peak_value = max(values) if values else 0 
        peak_index = values.index(peak_value) if peak_value > 0 and values else 0 
        peak_label = labels[peak_index] if labels else '' 
         
        return Response({ 
            "success": True, 
            "message": "Monthly usage chart data retrieved successfully", 
            "data": { 
                "labels": labels, 
                "values": values, 
                "total": round(total, 2), 
                "average": round(average, 2), 
                "unit": "litres", 
                "period": f"last_{months}_months", 
                "peak_value": round(peak_value, 2), 
                "peak_label": peak_label 
            } 
        }, status=status.HTTP_200_OK) 
         
    except Customer.DoesNotExist: 
        return Response({ 
            "success": False, 
            "message": "Customer not found" 
        }, status=status.HTTP_404_NOT_FOUND) 
    except Exception as e: 
        logger.error(f"Monthly chart data error: {str(e)}") 
        return Response({ 
            "success": False, 
            "message": f"Failed to retrieve chart data: {str(e)}" 
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
 
@api_view(['GET']) 
@csrf_exempt 
def customer_usage_hourly_chart(request, customer_id): 
    """ 
    Get hourly water usage data for chart visualization 
    GET: Returns chart-ready data for a specific day (default today) 
    Query params: date (YYYY-MM-DD, default today) 
    """ 
    try: 
        customer = Customer.objects.get(id=customer_id) 
         
        if not customer.block_unblock: 
            return Response({ 
                "success": False, 
                "message": "Account is blocked. Please contact support." 
            }, status=status.HTTP_403_FORBIDDEN) 
         
        # Get date from query params or use today 
        date_str = request.query_params.get('date') 
        if date_str: 
            try: 
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date() 
            except ValueError: 
                return Response({ 
                    "success": False, 
                    "message": "Invalid date format. Use YYYY-MM-DD" 
                }, status=status.HTTP_400_BAD_REQUEST) 
        else: 
            target_date = timezone.now().date() 
         
        labels = [] 
        values = [] 
         
        # Get usage for each hour (0-23) 
        for hour in range(24): 
            start = django_timezone.make_aware( 
                datetime.combine(target_date, datetime.min.time()) + timedelta(hours=hour) 
            ) 
            end = start + timedelta(hours=1) 
             
            hourly_usage = SensorData.objects.filter( 
                customer=customer, 
                esp_timestamp__range=(start, end) 
            ).aggregate(total=Sum('current_volume'))['total'] or 0 
             
            # Format label as 12-hour time 
            if hour == 0: 
                labels.append('12 AM') 
            elif hour < 12: 
                labels.append(f'{hour} AM') 
            elif hour == 12: 
                labels.append('12 PM') 
            else: 
                labels.append(f'{hour - 12} PM') 
             
            values.append(round(float(hourly_usage), 3)) 
         
        total = sum(values) 
        average = total / 24 
        peak_value = max(values) if values else 0 
        peak_index = values.index(peak_value) if peak_value > 0 and values else 0 
        peak_label = labels[peak_index] if labels else '' 
         
        return Response({ 
            "success": True, 
            "message": "Hourly usage chart data retrieved successfully", 
            "data": { 
                "labels": labels, 
                "values": values, 
                "total": round(total, 3), 
                "average": round(average, 3), 
                "unit": "litres", 
                "period": target_date.isoformat(), 
                "peak_value": round(peak_value, 3), 
                "peak_label": peak_label, 
                "date": target_date.isoformat() 
            } 
        }, status=status.HTTP_200_OK) 
         
    except Customer.DoesNotExist: 
        return Response({ 
            "success": False, 
            "message": "Customer not found" 
        }, status=status.HTTP_404_NOT_FOUND) 
    except Exception as e: 
        logger.error(f"Hourly chart data error: {str(e)}") 
        return Response({ 
            "success": False, 
            "message": f"Failed to retrieve chart data: {str(e)}" 
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
 
@api_view(['GET']) 
@csrf_exempt 
def customer_usage_comparison_chart(request, customer_id): 
    """ 
    Get usage comparison data (current vs previous period) 
    GET: Returns comparison data for charts 
    Query params: period (week/month, default week) 
    """ 
    try: 
        customer = Customer.objects.get(id=customer_id) 
         
        if not customer.block_unblock: 
            return Response({ 
                "success": False, 
                "message": "Account is blocked. Please contact support." 
            }, status=status.HTTP_403_FORBIDDEN) 
         
        period = request.query_params.get('period', 'week') 
        today = timezone.now().date() 
         
        if period == 'month': 
            # Current month vs previous month 
            current_start = today.replace(day=1) 
            prev_start = (current_start - timedelta(days=1)).replace(day=1) 
            prev_end = current_start - timedelta(days=1) 
            days_in_period = (today - current_start).days + 1 
            label_current = today.strftime('%B %Y') 
            label_previous = prev_start.strftime('%B %Y') 
        else: 
            # Current week vs previous week 
            days_in_period = 7 
            current_start = today - timedelta(days=6) 
            prev_start = current_start - timedelta(days=7) 
            prev_end = current_start - timedelta(days=1) 
            label_current = "This Week" 
            label_previous = "Last Week" 
         
        # Get current period usage 
        current_usage = WaterUsage.objects.filter( 
            customer=customer, 
            usage_date__gte=current_start, 
            usage_date__lte=today 
        ).aggregate(total=Sum('current_day_usage'))['total'] or 0 
         
        # Get previous period usage 
        previous_usage = WaterUsage.objects.filter( 
            customer=customer, 
            usage_date__gte=prev_start, 
            usage_date__lte=prev_end 
        ).aggregate(total=Sum('current_day_usage'))['total'] or 0 
         
        current_usage = float(current_usage) 
        previous_usage = float(previous_usage) 
         
        # Calculate change percentage 
        if previous_usage > 0: 
            change_percent = ((current_usage - previous_usage) / previous_usage) * 100 
        else: 
            change_percent = 100 if current_usage > 0 else 0 
         
        return Response({ 
            "success": True, 
            "message": "Usage comparison data retrieved successfully", 
            "data": { 
                "labels": [label_previous, label_current], 
                "values": [round(previous_usage, 2), round(current_usage, 2)], 
                "current_usage": round(current_usage, 2), 
                "previous_usage": round(previous_usage, 2), 
                "change_percent": round(change_percent, 1), 
                "change_direction": "up" if change_percent > 0 else "down" if 
change_percent < 0 else "same", 
                "unit": "litres", 
                "period": period 
            } 
        }, status=status.HTTP_200_OK) 
         
    except Customer.DoesNotExist: 
        return Response({ 
            "success": False, 
            "message": "Customer not found" 
        }, status=status.HTTP_404_NOT_FOUND) 
    except Exception as e: 
        logger.error(f"Comparison chart data error: {str(e)}") 
        return Response({ 
            "success": False, 
"message": f"Failed to retrieve comparison data: {str(e)}" 
}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ========== MOTIVATIONAL WATER QUOTES API ==========
@api_view(['GET'])
def water_quote_api(request):
    """
    API endpoint for fetching motivational water quotes.
    Can be used by Android apps, Streamlit applications, or any frontend.
    First checks database for custom quotes, falls back to default quotes.
    
    Returns:
        JSON response with a random motivational quote about water (with optional image).
    """
    from .models import MotivationalQuote
    
    # Try to get a quote from the database first
    db_quotes = MotivationalQuote.objects.filter(is_active=True)
    
    if db_quotes.exists():
        # Random quote from database
        selected_db_quote = random.choice(list(db_quotes))
        image_url = None
        if selected_db_quote.image:
            image_url = request.build_absolute_uri(selected_db_quote.image.url)
        
        return Response({
            "success": True,
            "quote": selected_db_quote.quote,
            "author": selected_db_quote.author,
            "image_url": image_url,
            "source": "database",
            "message": "Quote fetched successfully"
        }, status=status.HTTP_200_OK)
    
    # Fallback to default quotes
    water_quotes = [
        {"quote": "Water is the driving force of all nature.", "author": "Leonardo da Vinci"},
        {"quote": "Thousands have lived without love, not one without water.", "author": "W.H. Auden"},
        {"quote": "Water is life, and clean water means health.", "author": "Audrey Hepburn"},
        {"quote": "When the well is dry, we know the worth of water.", "author": "Benjamin Franklin"},
        {"quote": "In one drop of water are found all the secrets of all the oceans.", "author": "Kahlil Gibran"},
        {"quote": "Water is the softest thing, yet it can penetrate mountains and earth.", "author": "Lao Tzu"},
        {"quote": "We never know the worth of water till the well is dry.", "author": "Thomas Fuller"},
        {"quote": "Pure water is the world's first and foremost medicine.", "author": "Slovakian Proverb"},
        {"quote": "Water is the best of all things.", "author": "Pindar"},
        {"quote": "If there is magic on this planet, it is contained in water.", "author": "Loren Eiseley"},
        {"quote": "Nothing is softer or more flexible than water, yet nothing can resist it.", "author": "Lao Tzu"},
        {"quote": "Save water, and it will save you.", "author": "Unknown"},
        {"quote": "Water is the mirror that has the ability to show us what we cannot see.", "author": "Masaru Emoto"},
        {"quote": "The cure for anything is salt water: sweat, tears, or the sea.", "author": "Isak Dinesen"},
        {"quote": "A drop of water, if it could write out its own history, would explain the universe to us.", "author": "Lucy Larcom"},
        {"quote": "Water is sacred to all human beings.", "author": "Unknown"},
        {"quote": "When you drink the water, remember the spring.", "author": "Chinese Proverb"},
        {"quote": "Don't let the water run in the sink, our life's on the brink!", "author": "Unknown"},
        {"quote": "You don't drown by falling in water; you drown by staying there.", "author": "Edwin Louis Cole"},
        {"quote": "Water links us to our neighbor in a way more profound than any other.", "author": "John Thorson"},
        {"quote": "The wars of the twenty-first century will be fought over water.", "author": "Ismail Serageldin"},
        {"quote": "Conserve water, conserve life.", "author": "Unknown"},
        {"quote": "No water, no life. No blue, no green.", "author": "Sylvia Earle"},
        {"quote": "Water is precious. Waste not, want not.", "author": "Unknown"},
        {"quote": "Every drop of water counts. Make it count for good.", "author": "Unknown"},
    ]
    
    selected_quote = random.choice(water_quotes)
    
    return Response({
        "success": True,
        "quote": selected_quote["quote"],
        "author": selected_quote["author"],
        "image_url": None,
        "source": "default",
        "message": "Quote fetched successfully"
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def water_quotes_list_api(request):
    """
    API endpoint for fetching all motivational water quotes.
    Can be used by Android apps, Streamlit applications, or any frontend.
    
    Returns:
        JSON response with all motivational quotes about water.
    """
    from .models import MotivationalQuote
    
    db_quotes = MotivationalQuote.objects.filter(is_active=True)
    
    quotes_list = []
    for quote in db_quotes:
        image_url = None
        if quote.image:
            image_url = request.build_absolute_uri(quote.image.url)
        
        quotes_list.append({
            "id": quote.id,
            "quote": quote.quote,
            "author": quote.author,
            "image_url": image_url,
            "created_at": quote.created_at.isoformat()
        })
    
    return Response({
        "success": True,
        "count": len(quotes_list),
        "quotes": quotes_list,
        "message": "Quotes fetched successfully"
    }, status=status.HTTP_200_OK)





#================Create Razorpay Order API============================="""
@api_view(['POST'])
@csrf_exempt
def create_razorpay_order(request, customer_id):
    """
    Create Razorpay Order API for Android app
    POST: Create a Razorpay order before payment
    
    Required fields:
    - amount: Payment amount in rupees
    
    For subscription customers:
    - subscription_plan_id: ID of the subscription plan (optional, will use plan price)
    """
    import razorpay
    from django.conf import settings
    
    try:
        customer = Customer.objects.get(id=customer_id)
        
        amount = request.data.get('amount')
        subscription_plan_id = request.data.get('subscription_plan_id')
        
        # For subscription customers, get amount from plan if not provided
        if customer.customer_type == 'subscription' and subscription_plan_id and not amount:
            try:
                plan = SubscriptionPlan.objects.get(id=subscription_plan_id, status=True)
                amount = float(plan.price)
            except SubscriptionPlan.DoesNotExist:
                return Response({
                    "success": False,
                    "message": "Invalid subscription plan"
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if not amount:
            return Response({
                "success": False,
                "message": "Amount is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError("Amount must be positive")
        except (ValueError, TypeError):
            return Response({
                "success": False,
                "message": "Invalid amount value"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create Razorpay client
        client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
        
        # Create order (amount in paise)
        order_data = {
            'amount': int(amount * 100),  # Convert to paise
            'currency': 'INR',
            'receipt': f"rcpt_{uuid.uuid4().hex[:10]}",
            'notes': {
                'customer_id': str(customer.id),
                'customer_name': customer.name,
                'customer_email': customer.email,
                'customer_type': customer.customer_type
            }
        }
        
        razorpay_order = client.order.create(data=order_data)
        
        return Response({
            "success": True,
            "message": "Razorpay order created successfully",
            "data": {
                "razorpay_order_id": razorpay_order['id'],
                "razorpay_key_id": settings.RAZORPAY_KEY_ID,
                "amount": amount,
                "amount_in_paise": int(amount * 100),
                "currency": "INR",
                "customer": {
                    "id": str(customer.id),
                    "name": customer.name,
                    "email": customer.email,
                    "phone": customer.phone_number
                }
            }
        }, status=status.HTTP_200_OK)
        
    except Customer.DoesNotExist:
        return Response({
            "success": False,
            "message": "Customer not found"
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Create Razorpay order error: {str(e)}")
        return Response({
            "success": False,
            "message": f"Failed to create order: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

